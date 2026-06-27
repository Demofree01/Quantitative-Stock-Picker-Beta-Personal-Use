#!/usr/bin/env python3
from __future__ import annotations

import math
import os
import re
import sys
import subprocess
import time
import warnings
import json
import urllib.parse
import urllib.request
from concurrent.futures import ThreadPoolExecutor, as_completed, wait
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Tuple

import akshare as ak
import numpy as np
import pandas as pd
import yaml
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

warnings.filterwarnings(
    "ignore",
    message="urllib3 v2 only supports OpenSSL 1.1.1+",
    category=Warning,
)


PROJECT_DIR = Path(__file__).resolve().parent
CONFIG_PATH = PROJECT_DIR / "config.yaml"
HOLDINGS_PATH = PROJECT_DIR / "holdings.csv"
TRADE_RECORDS_PATH = PROJECT_DIR / "trades.csv"
TRADE_RECORDS_XLSX_PATH = PROJECT_DIR / "交割单.xlsx"
AUTO_HOLDINGS_PATH = PROJECT_DIR / "自动持仓.csv"
OUTPUT_DIR = PROJECT_DIR / "output"
CACHE_DIR = PROJECT_DIR / "cache"
HISTORY_CACHE_READ_ENABLED = True
RUN_STATUS_STEPS: List[Tuple[str, str]] = []

STOCK_PREFIXES = ("600", "601", "603", "605", "000", "001", "002", "003", "300", "301")
ETF_PREFIXES = ("510", "511", "512", "513", "515", "516", "517", "518", "520", "560", "561", "562", "563", "588", "589", "159")

ACTION_ORDER = {
    "建议卖出": 0,
    "建议减仓": 1,
    "观察": 2,
    "继续持有": 3,
    "可新开仓": 4,
}

ACTION_FILL_MAP = {
    "建议卖出": "FFF4CCCC",
    "建议减仓": "FFFCE5CD",
    "观察": "FFFFF2CC",
    "继续持有": "FFD9EAD3",
    "可新开仓": "FFD9EAF7",
}

CARD_FILL_MAP = {
    "default": "FFF8FAFC",
    "weak": "FFFFF2CC",
    "normal": "FFD9EAD3",
    "strong": "FFB6D7A8",
    "hot": "FF93C47D",
}

HEADER_FILL = PatternFill("solid", fgColor="FF1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", color="FFFFFFFF", bold=True, size=11)
TITLE_FILL = PatternFill("solid", fgColor="FF163A5F")
TITLE_FONT = Font(name="Microsoft YaHei", color="FFFFFFFF", bold=True, size=13)
BODY_FONT = Font(name="Microsoft YaHei", color="FF1F1F1F", size=10)
BODY_BOLD_FONT = Font(name="Microsoft YaHei", color="FF1F1F1F", bold=True, size=10)
CARD_FONT = Font(name="Microsoft YaHei", color="FF1F1F1F", bold=True, size=11)

THIN_SIDE = Side(style="thin", color="FFD9E2F3")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)


def load_config() -> Dict[str, Any]:
    if not CONFIG_PATH.exists():
        raise FileNotFoundError(f"未找到配置文件：{CONFIG_PATH}")
    with CONFIG_PATH.open("r", encoding="utf-8") as fh:
        data = yaml.safe_load(fh) or {}
    data.setdefault("universe", {})
    data.setdefault("valuation", {})
    data.setdefault("quality", {})
    data["universe"].setdefault("target_universe_size", int(data.get("data", {}).get("stock_candidate_size", 200)))
    data["universe"].setdefault("extra_rank_start", 200)
    data["universe"].setdefault("extra_rank_end", 220)
    data["universe"].setdefault("min_listing_days", 180)
    data["universe"].setdefault("min_price", float(data.get("filter", {}).get("min_price_stock", 3)))
    data["universe"].setdefault("min_avg_amount_20d", float(data.get("filter", {}).get("min_amount_stock", 80000000)))
    data["universe"].setdefault("min_float_market_cap", 5000000000)
    data["valuation"].setdefault("use_pe_ttm", True)
    data["valuation"].setdefault("pe_mode", "global_percentile_score")
    data["valuation"].setdefault("use_peg", False)
    data["quality"].setdefault("hard_filter", {})
    data["quality"]["hard_filter"].setdefault("roe_min", 0)
    data["quality"]["hard_filter"].setdefault("ocf_to_net_profit_min", 0)
    data["quality"]["hard_filter"].setdefault("deducted_profit_growth_min", -0.50)
    return data


def normalize_code(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    match = re.search(r"(\d{6})", text)
    return match.group(1) if match else ""


def clean_name(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.lower() == "nan":
        return ""
    return re.sub(r"\s+", "", text)


def is_st_like(name: Any) -> bool:
    text = clean_name(name).upper()
    risk_words = ["ST", "*ST", "退市", "风险警示"]
    return any(word in text for word in risk_words)


def infer_asset_type(code: str, raw: Any = "") -> str:
    raw_text = str(raw).strip().upper()
    if str(raw).strip() == "股票":
        return "STOCK"
    if raw_text in {"STOCK", "ETF", "CASH"}:
        return raw_text
    code = normalize_code(code)
    if code.startswith(ETF_PREFIXES):
        return "ETF"
    if code.startswith(STOCK_PREFIXES):
        return "STOCK"
    return raw_text if raw_text else ""


def is_cash_row(code: Any = "", name: Any = "", asset_type: Any = "") -> bool:
    text = " ".join([str(code), str(name), str(asset_type)]).strip().upper()
    return any(token in text for token in ["CASH", "现金", "可用资金", "剩余资金"])


def security_holdings(holdings: pd.DataFrame) -> pd.DataFrame:
    if holdings.empty or "asset_type" not in holdings.columns:
        return holdings.copy()
    return holdings[holdings["asset_type"].astype(str).str.upper() != "CASH"].copy()


def is_allowed_stock(code: str) -> bool:
    code = normalize_code(code)
    return len(code) == 6 and code.startswith(STOCK_PREFIXES)


def is_allowed_etf(code: str) -> bool:
    code = normalize_code(code)
    return len(code) == 6 and code.startswith(ETF_PREFIXES)


def symbol_for_stock(code: str) -> str:
    code = normalize_code(code)
    if code.startswith(("6",)):
        return f"sh{code}"
    if code.startswith(("0", "1", "2", "3")):
        return f"sz{code}"
    raise ValueError(f"不支持的股票代码：{code}")


def symbol_for_etf(code: str) -> str:
    code = normalize_code(code)
    if code.startswith("159"):
        return f"sz{code}"
    return f"sh{code}"


def to_number(value: Any) -> float:
    if value is None:
        return np.nan
    try:
        if pd.isna(value):
            return np.nan
    except Exception:
        pass
    try:
        return float(value)
    except Exception:
        text = str(value).replace(",", "").strip()
        if not text or text.lower() == "nan":
            return np.nan
        try:
            return float(text)
        except Exception:
            return np.nan


def to_int_or_none(value: Any) -> Optional[int]:
    num = to_number(value)
    if pd.isna(num):
        return None
    return int(round(num))


def to_share_int(value: Any) -> int:
    num = to_number(value)
    if pd.isna(num) or num <= 0:
        return 0
    return int(round(num))


def current_date_text() -> str:
    return current_date_obj().strftime("%Y%m%d")


def current_date_obj() -> datetime:
    override = os.environ.get("OPENCLAW_RUN_DATE", "").strip()
    if override:
        for fmt in ("%Y%m%d", "%Y-%m-%d"):
            try:
                parsed = datetime.strptime(override, fmt)
                now = datetime.now()
                return parsed.replace(hour=now.hour, minute=now.minute, second=now.second, microsecond=now.microsecond)
            except ValueError:
                pass
        raise ValueError(f"OPENCLAW_RUN_DATE 格式不正确，应为 YYYYMMDD 或 YYYY-MM-DD: {override}")
    return datetime.now()


def reset_status_steps() -> None:
    RUN_STATUS_STEPS.clear()


def add_status_step(name: str, status: str = "完成") -> None:
    RUN_STATUS_STEPS.append((name, status))


def show_status_window(title: str, finished: bool, output_path: Optional[Path] = None, error: Optional[BaseException] = None) -> None:
    if os.environ.get("OPENCLAW_HEADLESS", "").lower() in {"1", "true", "yes"} or not os.environ.get("DISPLAY"):
        print(title)
        for name, status in RUN_STATUS_STEPS:
            print(f"{status} - {name}")
        if output_path is not None:
            print(f"报告文件：{output_path}")
        if error is not None:
            print(f"错误：{error}")
        return
    try:
        import tkinter as tk
        from tkinter import ttk
    except Exception:
        print(title)
        for name, status in RUN_STATUS_STEPS:
            print(f"{status} - {name}")
        if output_path is not None:
            print(f"报告文件：{output_path}")
        if error is not None:
            print(f"错误：{error}")
        return

    try:
        root = tk.Tk()
    except Exception:
        print(title)
        for name, status in RUN_STATUS_STEPS:
            print(f"{status} - {name}")
        if output_path is not None:
            print(f"报告文件：{output_path}")
        if error is not None:
            print(f"错误：{error}")
        return
    root.title(title)
    root.geometry("560x420")
    root.resizable(False, False)

    main = ttk.Frame(root, padding=18)
    main.pack(fill="both", expand=True)

    status_text = "运行完成" if finished else "运行失败"
    status_color = "#1f7a3f" if finished else "#b42318"
    header = tk.Label(main, text=status_text, font=("Microsoft YaHei", 16, "bold"), fg=status_color)
    header.pack(anchor="w")

    if output_path is not None:
        tk.Label(main, text=f"报告文件：{output_path}", font=("Microsoft YaHei", 10), anchor="w", justify="left", wraplength=520).pack(anchor="w", pady=(8, 0))
    if error is not None:
        tk.Label(main, text=f"错误：{error}", font=("Microsoft YaHei", 10), fg="#b42318", anchor="w", justify="left", wraplength=520).pack(anchor="w", pady=(8, 0))

    ttk.Separator(main).pack(fill="x", pady=14)
    tk.Label(main, text="运行步骤", font=("Microsoft YaHei", 11, "bold")).pack(anchor="w")

    list_frame = ttk.Frame(main)
    list_frame.pack(fill="both", expand=True, pady=(8, 12))
    text = tk.Text(list_frame, height=12, wrap="word", font=("Microsoft YaHei", 10), relief="solid", borderwidth=1)
    text.pack(fill="both", expand=True)
    for idx, (name, status) in enumerate(RUN_STATUS_STEPS, start=1):
        text.insert("end", f"{idx}. [{status}] {name}\n")
    text.configure(state="disabled")

    ttk.Button(main, text="关闭", command=root.destroy).pack(anchor="e")
    root.mainloop()


def read_csv_with_encoding_fallback(path: Path, encodings: List[str]) -> pd.DataFrame:
    last_exc: Optional[Exception] = None
    for encoding in encodings:
        try:
            return pd.read_csv(path, dtype=str, encoding=encoding)
        except UnicodeDecodeError as exc:
            last_exc = exc
            continue
        except pd.errors.EmptyDataError:
            raise
        except Exception as exc:
            last_exc = exc
            continue
    if last_exc is not None:
        raise last_exc
    raise RuntimeError(f"无法读取CSV文件：{path}")


def safe_read_csv(path: Path) -> pd.DataFrame:
    required = ["code", "name", "asset_type", "shares", "cost_price"]
    if not path.exists():
        return pd.DataFrame(columns=required)
    try:
        df = read_csv_with_encoding_fallback(path, ["utf-8-sig", "utf-8", "gb18030", "gbk", "cp936"])
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=required)
    except Exception as exc:
        raise RuntimeError(
            f"读取持仓文件失败：{exc}。请确认 holdings.csv 是否为 UTF-8 或常见中文编码（GBK/GB18030）。"
        ) from exc
    for col in required:
        if col not in df.columns:
            df[col] = ""
    df = df[required].copy()
    df["name"] = df["name"].map(clean_name)
    cash_mask = [
        is_cash_row(code, name, raw_type)
        for code, name, raw_type in zip(df["code"].tolist(), df["name"].tolist(), df["asset_type"].tolist())
    ]
    df["code"] = [
        "CASH" if is_cash else normalize_code(code)
        for code, is_cash in zip(df["code"].tolist(), cash_mask)
    ]
    df["asset_type"] = [
        "CASH" if is_cash else infer_asset_type(code, raw_type)
        for code, raw_type, is_cash in zip(df["code"].tolist(), df["asset_type"].tolist(), cash_mask)
    ]
    df.loc[df["asset_type"] == "CASH", "name"] = df.loc[df["asset_type"] == "CASH", "name"].replace("", "现金")
    df["shares"] = pd.to_numeric(df["shares"], errors="coerce")
    df["cost_price"] = pd.to_numeric(df["cost_price"], errors="coerce")
    df.loc[df["asset_type"] == "CASH", "cost_price"] = df.loc[df["asset_type"] == "CASH", "cost_price"].fillna(1.0)
    df = df[df["code"] != ""].copy()
    df = df.drop_duplicates(subset=["asset_type", "code"], keep="first").reset_index(drop=True)
    return df


def normalize_trade_record_columns(df: pd.DataFrame) -> pd.DataFrame:
    rename_map = {
        "日期": "date",
        "交易日期": "date",
        "成交日期": "date",
        "方向": "action",
        "买卖方向": "action",
        "操作": "action",
        "代码": "code",
        "证券代码": "code",
        "名称": "name",
        "证券名称": "name",
        "标的类型": "asset_type",
        "类型": "asset_type",
        "成交数量": "shares",
        "数量": "shares",
        "股数": "shares",
        "成交价格": "price",
        "价格": "price",
        "成交价": "price",
        "手续费": "fee",
        "佣金": "fee",
        "印花税": "tax",
        "备注": "note",
    }
    out = df.copy()
    out.columns = [str(col).strip() for col in out.columns]
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    return out


def read_trade_records_file(path: Path) -> pd.DataFrame:
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return pd.read_excel(path, dtype=str)
    return read_csv_with_encoding_fallback(path, ["utf-8-sig", "utf-8", "gb18030", "gbk", "cp936"])


def resolve_trade_records_path() -> Optional[Path]:
    if TRADE_RECORDS_XLSX_PATH.exists():
        return TRADE_RECORDS_XLSX_PATH
    if TRADE_RECORDS_PATH.exists():
        return TRADE_RECORDS_PATH
    return None


def safe_read_trade_records(path: Optional[Path] = None) -> pd.DataFrame:
    required = ["date", "action", "code", "name", "asset_type", "shares", "price", "fee", "tax", "note"]
    path = path or resolve_trade_records_path()
    if path is None or not path.exists():
        return pd.DataFrame(columns=required)
    try:
        df = read_trade_records_file(path)
    except pd.errors.EmptyDataError:
        return pd.DataFrame(columns=required)
    except Exception as exc:
        raise RuntimeError(
            f"读取交易记录失败：{exc}。请确认交割单文件是否能正常打开。"
        ) from exc
    df = normalize_trade_record_columns(df)
    for col in required:
        if col not in df.columns:
            df[col] = ""
    df = df[required].copy()
    df["date"] = pd.to_datetime(df["date"], errors="coerce")
    df["code"] = df["code"].map(normalize_code)
    df["name"] = df["name"].map(clean_name)
    df["asset_type"] = [infer_asset_type(code, raw_type) for code, raw_type in zip(df["code"], df["asset_type"])]
    action_map = {
        "BUY": "BUY",
        "B": "BUY",
        "买入": "BUY",
        "开仓": "BUY",
        "SELL": "SELL",
        "S": "SELL",
        "卖出": "SELL",
        "平仓": "SELL",
        "减仓": "SELL",
    }
    df["action"] = df["action"].map(lambda x: action_map.get(str(x).strip().upper(), action_map.get(str(x).strip(), str(x).strip().upper())))
    df["shares"] = pd.to_numeric(df["shares"], errors="coerce").fillna(0.0)
    df["price"] = pd.to_numeric(df["price"], errors="coerce")
    df["fee"] = pd.to_numeric(df["fee"], errors="coerce").fillna(0.0)
    df["tax"] = pd.to_numeric(df["tax"], errors="coerce").fillna(0.0)
    df = df[(df["code"] != "") & (df["shares"] > 0) & (df["price"].notna()) & (df["price"] > 0)].copy()
    df = df[df["action"].isin(["BUY", "SELL"])].copy()
    df = df.sort_values(["date", "code", "action"], na_position="last").reset_index(drop=True)
    return df


def seed_lots_from_holdings(holdings: pd.DataFrame) -> Dict[Tuple[str, str], List[Dict[str, float]]]:
    lots: Dict[Tuple[str, str], List[Dict[str, float]]] = {}
    if holdings.empty:
        return lots
    for _, row in security_holdings(holdings).iterrows():
        code = normalize_code(row.get("code"))
        asset_type = infer_asset_type(code, row.get("asset_type"))
        shares = to_number(row.get("shares"))
        cost_price = to_number(row.get("cost_price"))
        if not code or pd.isna(shares) or shares <= 0 or pd.isna(cost_price) or cost_price <= 0:
            continue
        lots.setdefault((asset_type, code), []).append({"shares": float(shares), "unit_cost": float(cost_price)})
    return lots


def build_realized_pnl_from_trades(trades: pd.DataFrame, starting_holdings: Optional[pd.DataFrame] = None) -> Tuple[Dict[str, float], pd.DataFrame, float, float]:
    if trades.empty:
        return {"realized_pnl": 0.0, "total_buy_amount": 0.0, "total_sell_amount": 0.0, "unmatched_sell_shares": 0.0}, pd.DataFrame(), 0.0, 0.0
    lots = seed_lots_from_holdings(starting_holdings if starting_holdings is not None else pd.DataFrame())
    records: List[Dict[str, Any]] = []
    total_buy_amount = 0.0
    total_sell_amount = 0.0
    realized_pnl = 0.0
    unmatched_sell_shares = 0.0
    for _, row in trades.iterrows():
        key = (str(row["asset_type"]), str(row["code"]))
        action = str(row["action"])
        shares = float(row["shares"])
        price = float(row["price"])
        fee = float(row.get("fee", 0.0) or 0.0)
        tax = float(row.get("tax", 0.0) or 0.0)
        gross_amount = shares * price
        if action == "BUY":
            total_buy_amount += gross_amount + fee + tax
            unit_cost = (gross_amount + fee + tax) / shares if shares > 0 else price
            lots.setdefault(key, []).append({"shares": shares, "unit_cost": unit_cost})
            records.append(
                {
                    "date": row["date"],
                    "action": "买入",
                    "code": row["code"],
                    "name": row["name"],
                    "asset_type": chinese_asset_type(row["asset_type"]),
                    "shares": shares,
                    "price": price,
                    "amount": gross_amount,
                    "cost_basis": np.nan,
                    "realized_pnl": np.nan,
                    "note": row.get("note", ""),
                }
            )
            continue

        total_sell_amount += gross_amount - fee - tax
        remaining = shares
        cost_basis = 0.0
        for lot in lots.setdefault(key, []):
            if remaining <= 0:
                break
            lot_shares = float(lot.get("shares", 0.0))
            if lot_shares <= 0:
                continue
            matched = min(remaining, lot_shares)
            cost_basis += matched * float(lot["unit_cost"])
            lot["shares"] = lot_shares - matched
            remaining -= matched
        lots[key] = [lot for lot in lots.get(key, []) if float(lot.get("shares", 0.0)) > 1e-8]
        unmatched_sell_shares += max(0.0, remaining)
        net_sell_amount = gross_amount - fee - tax
        pnl = net_sell_amount - cost_basis if remaining < shares else np.nan
        if pd.notna(pnl):
            realized_pnl += float(pnl)
        records.append(
            {
                "date": row["date"],
                "action": "卖出",
                "code": row["code"],
                "name": row["name"],
                "asset_type": chinese_asset_type(row["asset_type"]),
                "shares": shares,
                "price": price,
                "amount": gross_amount,
                "cost_basis": cost_basis if remaining < shares else np.nan,
                "realized_pnl": pnl,
                "note": row.get("note", ""),
            }
        )
    summary = {
        "realized_pnl": float(realized_pnl),
        "total_buy_amount": float(total_buy_amount),
        "total_sell_amount": float(total_sell_amount),
        "unmatched_sell_shares": float(unmatched_sell_shares),
    }
    detail = pd.DataFrame(records)
    if not detail.empty:
        detail = detail.rename(
            columns={
                "date": "日期",
                "action": "方向",
                "code": "代码",
                "name": "名称",
                "asset_type": "标的类型",
                "shares": "成交数量",
                "price": "成交价格",
                "amount": "成交金额",
                "cost_basis": "卖出匹配成本",
                "realized_pnl": "已实现盈亏",
                "note": "备注",
            }
        )
    return summary, detail, total_buy_amount, total_sell_amount


def apply_trade_records_to_holdings(base_holdings: pd.DataFrame, trades: pd.DataFrame) -> pd.DataFrame:
    required = ["code", "name", "asset_type", "shares", "cost_price"]
    if base_holdings.empty:
        holdings_map: Dict[Tuple[str, str], Dict[str, Any]] = {}
        cash = 0.0
    else:
        holdings_map = {}
        cash = cash_balance(base_holdings)
        for _, row in security_holdings(base_holdings).iterrows():
            code = normalize_code(row.get("code"))
            asset_type = infer_asset_type(code, row.get("asset_type"))
            if not code or asset_type not in {"STOCK", "ETF"}:
                continue
            shares = to_number(row.get("shares"))
            cost_price = to_number(row.get("cost_price"))
            if pd.isna(shares) or shares <= 0:
                continue
            holdings_map[(asset_type, code)] = {
                "code": code,
                "name": clean_name(row.get("name")),
                "asset_type": asset_type,
                "shares": float(shares),
                "cost_price": float(cost_price) if pd.notna(cost_price) and cost_price > 0 else 0.0,
            }

    for _, row in trades.iterrows():
        code = normalize_code(row.get("code"))
        asset_type = infer_asset_type(code, row.get("asset_type"))
        if not code or asset_type not in {"STOCK", "ETF"}:
            continue
        action = str(row.get("action", "")).upper()
        shares = float(row.get("shares", 0.0) or 0.0)
        price = float(row.get("price", 0.0) or 0.0)
        fee = float(row.get("fee", 0.0) or 0.0)
        tax = float(row.get("tax", 0.0) or 0.0)
        if shares <= 0 or price <= 0:
            continue
        key = (asset_type, code)
        gross_amount = shares * price
        if action == "BUY":
            cash -= gross_amount + fee + tax
            current = holdings_map.get(
                key,
                {"code": code, "name": clean_name(row.get("name")), "asset_type": asset_type, "shares": 0.0, "cost_price": 0.0},
            )
            old_shares = float(current.get("shares", 0.0) or 0.0)
            old_cost = float(current.get("cost_price", 0.0) or 0.0)
            new_shares = old_shares + shares
            new_cost_total = old_shares * old_cost + gross_amount + fee + tax
            current["shares"] = new_shares
            current["cost_price"] = new_cost_total / new_shares if new_shares > 0 else 0.0
            if clean_name(row.get("name")):
                current["name"] = clean_name(row.get("name"))
            holdings_map[key] = current
        elif action == "SELL":
            cash += gross_amount - fee - tax
            current = holdings_map.get(
                key,
                {"code": code, "name": clean_name(row.get("name")), "asset_type": asset_type, "shares": 0.0, "cost_price": price},
            )
            current["shares"] = max(0.0, float(current.get("shares", 0.0) or 0.0) - shares)
            if clean_name(row.get("name")):
                current["name"] = clean_name(row.get("name"))
            if current["shares"] > 1e-8:
                holdings_map[key] = current
            elif key in holdings_map:
                del holdings_map[key]

    rows = list(holdings_map.values())
    rows = sorted(rows, key=lambda r: (str(r["asset_type"]), str(r["code"])))
    rows.append({"code": "CASH", "name": "现金", "asset_type": "CASH", "shares": cash, "cost_price": 1.0})
    out = pd.DataFrame(rows, columns=required)
    out["shares"] = pd.to_numeric(out["shares"], errors="coerce").round(4)
    out["cost_price"] = pd.to_numeric(out["cost_price"], errors="coerce").round(6)
    return out


def suppress_output() -> Any:
    from contextlib import redirect_stderr, redirect_stdout
    import io

    return redirect_stdout(io.StringIO()), redirect_stderr(io.StringIO())


def call_with_retry(func, *args, retries: int = 1, pause: float = 0.6, **kwargs):
    last_exc = None
    for attempt in range(retries + 1):
        try:
            return func(*args, **kwargs)
        except Exception as exc:  # pragma: no cover - network dependent
            last_exc = exc
            if attempt < retries:
                time.sleep(pause * (attempt + 1))
    raise last_exc


def normalize_spot_frame(df: pd.DataFrame, asset_type: str) -> pd.DataFrame:
    out = df.copy()
    rename_map = {
        "code": "代码",
        "name": "名称",
        "latest_price": "最新价",
        "spot_amount": "成交额",
        "volume": "成交量",
        "spot_volume": "成交量",
        "market_cap": "总市值",
        "total_market_cap": "总市值",
        "spot_market_cap": "总市值",
        "float_market_cap": "流通市值",
        "上市日期": "上市日期",
        "listing_date": "上市日期",
        "industry": "行业",
        "所属行业": "行业",
        "pe_ttm": "PE_TTM",
        "peTTM": "PE_TTM",
        "PE_TTM": "PE_TTM",
        "peg": "PEG",
        "PEG": "PEG",
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    for col in ["代码", "名称", "最新价", "成交额"]:
        if col not in out.columns:
            out[col] = ""
    if "成交量" not in out.columns:
        out["成交量"] = np.nan
    if "总市值" not in out.columns:
        out["总市值"] = np.nan
    if "流通市值" not in out.columns:
        out["流通市值"] = np.nan
    if "上市日期" not in out.columns:
        out["上市日期"] = pd.NaT
    if "行业" not in out.columns:
        out["行业"] = "未知"
    if "PE_TTM" not in out.columns:
        pe_aliases = ["市盈率-TTM", "市盈率(TTM)", "市盈率-动态", "动态市盈率"]
        out["PE_TTM"] = next((out[col] for col in pe_aliases if col in out.columns), np.nan)
    if "PEG" not in out.columns:
        out["PEG"] = np.nan
    if out["总市值"].isna().all() and "流通市值" in out.columns:
        out["总市值"] = out["流通市值"]
    out["代码"] = out["代码"].map(normalize_code)
    out["名称"] = out["名称"].map(clean_name)
    for col in ["最新价", "成交额", "成交量", "总市值", "流通市值", "PE_TTM", "PEG"]:
        out[col] = pd.to_numeric(out[col], errors="coerce")
    out["上市日期"] = pd.to_datetime(out["上市日期"], errors="coerce")
    out["行业"] = out["行业"].fillna("未知").astype(str).str.strip().replace("", "未知")
    out["asset_type"] = asset_type
    out = out[out["代码"] != ""].copy()
    return out[["代码", "名称", "最新价", "成交额", "成交量", "总市值", "流通市值", "上市日期", "行业", "PE_TTM", "PEG", "asset_type"]].copy()


def download_stock_spot_eastmoney() -> pd.DataFrame:
    """Fetch A-share spot data directly from Eastmoney push2."""
    fields = "f12,f14,f2,f5,f6,f20,f21,f9,f26,f100"
    hosts = [
        "https://push2.eastmoney.com/api/qt/clist/get",
        "https://82.push2.eastmoney.com/api/qt/clist/get",
        "https://75.push2.eastmoney.com/api/qt/clist/get",
    ]
    params = {
        "pn": 1,
        "pz": 6000,
        "po": 1,
        "np": 1,
        "ut": "bd1d9ddb04089700cf9c27f6f7426281",
        "fltt": 2,
        "invt": 2,
        "fid": "f6",
        "fs": "m:0+t:6,m:0+t:80,m:1+t:2,m:1+t:23",
        "fields": fields,
    }
    last_exc: Optional[Exception] = None
    for base_url in hosts:
        try:
            url = base_url + "?" + urllib.parse.urlencode(params)
            req = urllib.request.Request(
                url,
                headers={
                    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
                    "Referer": "https://quote.eastmoney.com/",
                    "Accept": "application/json,text/plain,*/*",
                },
            )
            with urllib.request.urlopen(req, timeout=30) as resp:
                payload = json.loads(resp.read().decode("utf-8"))
            rows = (((payload or {}).get("data") or {}).get("diff") or [])
            if not rows:
                raise RuntimeError("东方财富快照接口返回为空")
            out = pd.DataFrame(rows).rename(
                columns={
                    "f12": "代码",
                    "f14": "名称",
                    "f2": "最新价",
                    "f5": "成交量",
                    "f6": "成交额",
                    "f20": "总市值",
                    "f21": "流通市值",
                    "f26": "上市日期",
                    "f100": "行业",
                    "f9": "PE_TTM",
                }
            )
            out["PEG"] = np.nan
            return normalize_spot_frame(out, "STOCK")
        except Exception as exc:
            last_exc = exc
    if last_exc:
        raise last_exc
    raise RuntimeError("东方财富快照接口失败")


def download_stock_spot() -> pd.DataFrame:
    last_exc: Optional[Exception] = None
    for func in [download_stock_spot_eastmoney, getattr(ak, "stock_zh_a_spot_em", None), ak.stock_zh_a_spot]:
        if func is None:
            continue
        try:
            return normalize_spot_frame(call_with_retry(func, retries=1), "STOCK")
        except Exception as exc:
            last_exc = exc
    if last_exc is not None:
        raise last_exc
    raise RuntimeError("股票现货接口失败。")


def download_stock_code_list() -> pd.DataFrame:
    df = call_with_retry(ak.stock_info_a_code_name, retries=1)
    df = df.copy()
    df["代码"] = df["code"].map(normalize_code)
    df["名称"] = df["name"].map(clean_name)
    return df[["代码", "名称"]].drop_duplicates(subset=["代码"]).reset_index(drop=True)


def download_etf_spot() -> pd.DataFrame:
    return normalize_spot_frame(call_with_retry(ak.fund_etf_spot_em, retries=1), "ETF")


def symbol_for_em(code: str) -> str:
    code = normalize_code(code)
    if code.startswith("6"):
        return f"SH{code}"
    return f"SZ{code}"


def fetch_stock_valuation_metrics(code: str) -> Dict[str, float]:
    metrics = {"PE_TTM": np.nan, "PEG": np.nan, "MARKET_CAP": np.nan}
    stdout_ctx, stderr_ctx = suppress_output()
    try:
        with stdout_ctx, stderr_ctx:
            df = call_with_retry(ak.stock_zh_valuation_comparison_em, symbol=symbol_for_em(code), retries=1)
        if not df.empty:
            rows = df[df["代码"].astype(str).map(normalize_code) == normalize_code(code)] if "代码" in df.columns else pd.DataFrame()
            # stock_zh_valuation_comparison_em sometimes returns only peer/industry rows.
            # Do not borrow a peer's PE for the target stock; fall back to Baidu valuation instead.
            if not rows.empty:
                row = rows.iloc[0]
                metrics["PE_TTM"] = to_number(row.get("市盈率-TTM"))
                metrics["PEG"] = to_number(row.get("PEG"))
    except Exception:
        pass
    if pd.isna(metrics["PE_TTM"]):
        try:
            stdout_ctx, stderr_ctx = suppress_output()
            with stdout_ctx, stderr_ctx:
                pe_df = call_with_retry(ak.stock_zh_valuation_baidu, symbol=normalize_code(code), indicator="市盈率(TTM)", period="近一年", retries=1)
            if not pe_df.empty:
                latest_pe = to_number(pe_df["value"].dropna().iloc[-1])
                if pd.notna(latest_pe):
                    metrics["PE_TTM"] = latest_pe
        except Exception:
            pass
    try:
        stdout_ctx, stderr_ctx = suppress_output()
        with stdout_ctx, stderr_ctx:
            cap_df = call_with_retry(ak.stock_zh_valuation_baidu, symbol=normalize_code(code), indicator="总市值", period="近一年", retries=1)
        if not cap_df.empty:
            latest_cap = to_number(cap_df["value"].dropna().iloc[-1])
            if pd.notna(latest_cap):
                metrics["MARKET_CAP"] = latest_cap * 100000000.0
    except Exception:
        pass
    return metrics


def populate_stock_fundamental_metrics(df: pd.DataFrame, cfg: Dict[str, Any]) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    out = df.copy()
    for col in ["PE_TTM", "PEG", "总市值"]:
        if col not in out.columns:
            out[col] = np.nan
    if bool(cfg["data"].get("disable_slow_fundamental_fetch", True)):
        return out
    need_fetch = out["PE_TTM"].isna() | out["总市值"].isna()
    if bool(cfg.get("valuation", {}).get("use_peg", False)):
        need_fetch = need_fetch | out["PEG"].isna()
    if need_fetch.any():
        use_subprocess = bool(cfg["data"].get("use_subprocess_fundamental_fetch", True))
        per_symbol_timeout = float(cfg["data"].get("fundamental_per_symbol_timeout_seconds", 8))
        fetch_items = list(out.loc[need_fetch, "代码"].items())
        fetch_limit = int(cfg["data"].get("fundamental_prefetch_limit", len(fetch_items)) or len(fetch_items))
        if len(fetch_items) > fetch_limit:
            print(f"估值补取数量限制：{len(fetch_items)} -> {fetch_limit}，其余PE缺失按估值得分容错处理。")
            fetch_items = fetch_items[:fetch_limit]
        if use_subprocess:
            total = len(fetch_items)
            for n, (idx, code) in enumerate(fetch_items, 1):
                metrics = fetch_stock_valuation_metrics_subprocess(code, timeout_seconds=per_symbol_timeout)
                if pd.isna(out.at[idx, "PE_TTM"]):
                    out.at[idx, "PE_TTM"] = metrics.get("PE_TTM", np.nan)
                if pd.isna(out.at[idx, "PEG"]):
                    out.at[idx, "PEG"] = metrics.get("PEG", np.nan)
                if pd.isna(out.at[idx, "总市值"]):
                    out.at[idx, "总市值"] = metrics.get("MARKET_CAP", np.nan)
                if n == 1 or n % 10 == 0 or n == total:
                    print(f"估值补取进度：{n}/{total}", flush=True)
        else:
            workers = min(int(cfg["data"].get("max_workers", 12)), max(1, int(need_fetch.sum())))
            timeout_seconds = float(cfg["data"].get("fundamental_fetch_timeout_seconds", 75))
            executor = ThreadPoolExecutor(max_workers=workers)
            future_map = {
                executor.submit(fetch_stock_valuation_metrics, code): idx
                for idx, code in fetch_items
            }
            done, not_done = wait(list(future_map.keys()), timeout=timeout_seconds)
            for future in done:
                idx = future_map[future]
                try:
                    metrics = future.result()
                except Exception:
                    metrics = {"PE_TTM": np.nan, "PEG": np.nan, "MARKET_CAP": np.nan}
                if pd.isna(out.at[idx, "PE_TTM"]):
                    out.at[idx, "PE_TTM"] = metrics.get("PE_TTM", np.nan)
                if pd.isna(out.at[idx, "PEG"]):
                    out.at[idx, "PEG"] = metrics.get("PEG", np.nan)
                if pd.isna(out.at[idx, "总市值"]):
                    out.at[idx, "总市值"] = metrics.get("MARKET_CAP", np.nan)
            for future in not_done:
                future.cancel()
            if not_done:
                print(f"估值补取超时：{len(not_done)} 只股票未完成，已跳过这些估值字段。")
            executor.shutdown(wait=False, cancel_futures=True)
    return out


def fetch_stock_valuation_metrics_subprocess(code: str, timeout_seconds: float = 12.0) -> Dict[str, float]:
    """Fetch PE/PEG/market cap in a child process with a hard timeout."""
    child_code = """
import json, re, sys, warnings
warnings.filterwarnings('ignore')
import akshare as ak
import numpy as np

def normalize_code(value):
    text = str(value or '').strip()
    m = re.search(r'(\\d{6})', text)
    return m.group(1) if m else ''

def to_number(value):
    try:
        if value is None:
            return float('nan')
        return float(str(value).replace(',', '').strip())
    except Exception:
        return float('nan')

def symbol_for_em(code):
    code = normalize_code(code)
    return ('SH' if code.startswith('6') else 'SZ') + code

code = normalize_code(sys.argv[1])
metrics = {'PE_TTM': float('nan'), 'PEG': float('nan'), 'MARKET_CAP': float('nan')}
try:
    df = ak.stock_zh_valuation_comparison_em(symbol=symbol_for_em(code))
    if df is not None and not df.empty:
        rows = df[df['代码'].astype(str).map(normalize_code) == code] if '代码' in df.columns else None
        if rows is not None and not rows.empty:
            row = rows.iloc[0]
            metrics['PE_TTM'] = to_number(row.get('市盈率-TTM'))
            metrics['PEG'] = to_number(row.get('PEG'))
except Exception:
    pass
try:
    if not np.isfinite(metrics['PE_TTM']):
        pe_df = ak.stock_zh_valuation_baidu(symbol=code, indicator='市盈率(TTM)', period='近一年')
        if pe_df is not None and not pe_df.empty:
            metrics['PE_TTM'] = to_number(pe_df['value'].dropna().iloc[-1])
except Exception:
    pass
try:
    cap_df = ak.stock_zh_valuation_baidu(symbol=code, indicator='总市值', period='近一年')
    if cap_df is not None and not cap_df.empty:
        cap = to_number(cap_df['value'].dropna().iloc[-1])
        if np.isfinite(cap):
            metrics['MARKET_CAP'] = cap * 100000000.0
except Exception:
    pass
print(json.dumps(metrics, ensure_ascii=False, allow_nan=True))
"""
    try:
        proc = subprocess.run(
            [sys.executable, '-c', child_code, normalize_code(code)],
            text=True,
            capture_output=True,
            timeout=timeout_seconds,
        )
        if proc.returncode != 0:
            raise RuntimeError((proc.stderr or proc.stdout or '子进程估值接口失败').strip())
        lines = [line.strip() for line in proc.stdout.splitlines() if line.strip()]
        if not lines:
            raise RuntimeError('子进程估值接口无输出')
        data = json.loads(lines[-1])
        return {
            'PE_TTM': to_number(data.get('PE_TTM')),
            'PEG': to_number(data.get('PEG')),
            'MARKET_CAP': to_number(data.get('MARKET_CAP')),
        }
    except subprocess.TimeoutExpired:
        return {'PE_TTM': np.nan, 'PEG': np.nan, 'MARKET_CAP': np.nan}
    except Exception:
        return {'PE_TTM': np.nan, 'PEG': np.nan, 'MARKET_CAP': np.nan}


def apply_stock_fundamental_filters(df: pd.DataFrame, cfg: Dict[str, Any]) -> pd.DataFrame:
    """Keep valuation data, but do not hard-filter PE/PEG under the simplified model."""
    if df.empty:
        return df
    out = populate_stock_fundamental_metrics(df, cfg)
    for col in ["PE_TTM", "PEG", "总市值", "流通市值", "ROE", "经营现金流/净利润", "扣非净利润同比"]:
        if col not in out.columns:
            out[col] = np.nan
        out[col] = pd.to_numeric(out[col], errors="coerce")
    qcfg = cfg.get("quality", {}).get("hard_filter", {})
    roe_min = float(qcfg.get("roe_min", 0))
    ocf_min = float(qcfg.get("ocf_to_net_profit_min", 0))
    growth_min = float(qcfg.get("deducted_profit_growth_min", -0.50))
    # 财务字段缺失时先容错放行；有值才做排雷硬剔除。
    out = out[out["ROE"].isna() | (out["ROE"] >= roe_min)].copy()
    out = out[out["经营现金流/净利润"].isna() | (out["经营现金流/净利润"] >= ocf_min)].copy()
    out = out[out["扣非净利润同比"].isna() | (out["扣非净利润同比"] >= growth_min)].copy()
    return out


def bounded_score(series: pd.Series, pass_line: float, good_line: float, neutral: float = 0.5) -> pd.Series:
    values = pd.to_numeric(series, errors="coerce")
    if good_line == pass_line:
        score = pd.Series(neutral, index=values.index, dtype=float)
    else:
        score = (values - pass_line) / (good_line - pass_line)
    return score.clip(0, 1).fillna(neutral) * 100.0


def compute_valuation_score(df: pd.DataFrame, cfg: Dict[str, Any]) -> pd.Series:
    if not bool(cfg.get("valuation", {}).get("use_pe_ttm", True)):
        return pd.Series(50.0, index=df.index)
    pe = pd.to_numeric(df.get("pe_ttm", df.get("PE_TTM", pd.Series(np.nan, index=df.index))), errors="coerce")
    score = pd.Series(30.0, index=df.index, dtype=float)
    valid = pe > 0
    if valid.any():
        # 全样本 PE_TTM 反向分位：候选池内 PE 越低，估值得分越高。
        # 当前数据源的行业字段不稳定，因此不再按行业分组。
        score.loc[valid] = pct_rank(pe, higher_is_better=False).loc[valid]
    score.loc[pe > 80] = score.loc[pe > 80].clip(upper=35.0)
    score.loc[pe.isna() | (pe <= 0)] = 30.0
    return score.fillna(30.0)



def listing_days_from_series(series: pd.Series) -> pd.Series:
    dates = pd.to_datetime(series, errors="coerce")
    today = pd.Timestamp(current_date_obj().date())
    return (today - dates).dt.days


def build_stock_candidates(cfg: Dict[str, Any]) -> pd.DataFrame:
    stock_spot = download_stock_spot()
    keep_cols = ["代码", "名称", "最新价", "成交额", "成交量", "总市值", "流通市值", "上市日期", "行业", "PE_TTM", "PEG"]
    for col in keep_cols:
        if col not in stock_spot.columns:
            stock_spot[col] = np.nan
    merged = stock_spot[keep_cols].copy()
    raw_count = len(merged)
    universe_cfg = cfg.get("universe", {})
    target_size = int(universe_cfg.get("target_universe_size", cfg.get("data", {}).get("stock_candidate_size", 200)))
    extra_rank_start = int(universe_cfg.get("extra_rank_start", 200))
    extra_rank_end = int(universe_cfg.get("extra_rank_end", 220))
    min_listing_days = int(universe_cfg.get("min_listing_days", 180))
    min_price = float(universe_cfg.get("min_price", cfg.get("filter", {}).get("min_price_stock", 3)))
    min_amount = float(universe_cfg.get("min_avg_amount_20d", cfg.get("filter", {}).get("min_amount_stock", 80000000)))
    min_float_cap = float(universe_cfg.get("min_float_market_cap", 5000000000))

    merged["名称"] = merged["名称"].map(clean_name)
    merged["asset_type"] = "STOCK"
    merged = merged[merged["代码"].str.startswith(STOCK_PREFIXES)].copy()
    merged = merged[~merged["名称"].map(is_st_like)].copy()
    merged["最新价"] = pd.to_numeric(merged["最新价"], errors="coerce")
    merged["成交额"] = pd.to_numeric(merged["成交额"], errors="coerce")
    merged["总市值"] = pd.to_numeric(merged["总市值"], errors="coerce")
    merged["流通市值"] = pd.to_numeric(merged["流通市值"], errors="coerce")
    merged["上市天数"] = listing_days_from_series(merged["上市日期"])
    merged["流通市值"] = merged["流通市值"].fillna(merged["总市值"])
    has_float_cap = merged["流通市值"].notna().any()

    merged = merged[merged["最新价"] >= min_price].copy()
    merged = merged[merged["成交额"] >= min_amount].copy()
    if has_float_cap:
        merged = merged[merged["流通市值"] >= min_float_cap].copy()
    else:
        print("股票快照缺少流通市值/总市值字段，已跳过50亿元流通市值硬过滤，先按成交额排名容错运行。")
    # 上市日期缺失时放行，避免数据源字段缺失导致整批股票被误删；有日期则严格执行 180 天。
    merged = merged[merged["上市天数"].isna() | (merged["上市天数"] >= min_listing_days)].copy()

    if merged.empty:
        print(f"股票快照初筛：{raw_count} -> 0；股票池为空。")
        for col in ["股票池排名分", "流通市值排名分", "成交额排名分"]:
            merged[col] = np.nan
        return merged

    if has_float_cap:
        merged["流通市值排名分"] = pct_rank(merged["流通市值"], higher_is_better=True)
    else:
        merged["流通市值排名分"] = 50.0
    merged["成交额排名分"] = pct_rank(merged["成交额"], higher_is_better=True)
    merged["股票池排名分"] = merged["流通市值排名分"] * 0.60 + merged["成交额排名分"] * 0.40
    merged = merged.sort_values(["股票池排名分", "流通市值", "成交额"], ascending=[False, False, False]).reset_index(drop=True)
    merged["股票池初筛排名"] = np.arange(1, len(merged) + 1)
    top_pool = merged.head(target_size)
    if extra_rank_start > 0 and extra_rank_end >= extra_rank_start:
        extra_pool = merged[(merged["股票池初筛排名"] >= extra_rank_start) & (merged["股票池初筛排名"] <= extra_rank_end)]
        merged = pd.concat([top_pool, extra_pool], ignore_index=True).drop_duplicates(subset=["代码"], keep="first")
    else:
        merged = top_pool.copy()

    # 只对最终股票池补 PE/市值，PEG 按新策略禁用，不参与选股。
    merged = populate_stock_fundamental_metrics(merged, cfg).reset_index(drop=True)
    snapshot_count = len(merged)
    print(f"股票池初筛：{raw_count} -> {snapshot_count}；使用流通市值+成交额综合排名，保留前{target_size}只 + 第{extra_rank_start}-{extra_rank_end}名。")
    return merged[[
        "代码", "名称", "asset_type", "最新价", "成交额", "成交量", "总市值", "流通市值", "上市日期", "上市天数", "行业",
        "股票池排名分", "股票池初筛排名", "流通市值排名分", "成交额排名分", "PE_TTM", "PEG",
    ]].copy()


def history_cache_path(asset_type: str, code: str, cfg: Dict[str, Any]) -> Path:
    adjust = str(cfg["data"].get("adjust", "qfq"))
    return CACHE_DIR / f"{asset_type.lower()}_{normalize_code(code)}_{adjust}_{current_date_text()}.csv"


def successful_run_marker_path() -> Path:
    return CACHE_DIR / f"successful_run_{current_date_text()}.marker"


def read_history_cache(asset_type: str, code: str, cfg: Dict[str, Any]) -> Optional[pd.DataFrame]:
    if not bool(cfg["data"].get("cache_history", True)):
        return None
    if not HISTORY_CACHE_READ_ENABLED:
        return None
    path = history_cache_path(asset_type, code, cfg)
    if not path.exists():
        # 周末/非交易日运行时，允许复用同代码最近一份历史行情缓存，
        # 避免为了验证周报逻辑反复触发大量联网请求。
        adjust = str(cfg["data"].get("adjust", "qfq"))
        candidates = sorted(CACHE_DIR.glob(f"{asset_type.lower()}_{normalize_code(code)}_{adjust}_*.csv"), reverse=True)
        if not candidates:
            return None
        path = candidates[0]
    try:
        df = pd.read_csv(path)
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        for col in ["open", "close", "high", "low", "amount", "turnover_amount"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        df = df.dropna(subset=["date", "close", "amount"]).sort_values("date").reset_index(drop=True)
        if len(df) >= int(cfg["filter"]["min_history_bars"]):
            return df[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()
    except Exception:
        return None
    return None


def write_history_cache(asset_type: str, code: str, cfg: Dict[str, Any], df: pd.DataFrame) -> None:
    if not bool(cfg["data"].get("cache_history", True)):
        return
    try:
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        df.to_csv(history_cache_path(asset_type, code, cfg), index=False, encoding="utf-8-sig")
    except Exception:
        pass


def build_etf_candidates(cfg: Dict[str, Any]) -> pd.DataFrame:
    etf_spot = download_etf_spot()
    merged = etf_spot.copy()
    merged["asset_type"] = "ETF"
    merged = merged[merged["代码"].str.startswith(ETF_PREFIXES)].copy()
    merged = merged[~merged["名称"].map(is_st_like)].copy()
    merged = merged[pd.to_numeric(merged["最新价"], errors="coerce") >= float(cfg["filter"]["min_price_etf"])].copy()
    merged = merged[pd.to_numeric(merged["成交额"], errors="coerce") >= float(cfg["filter"]["min_amount_etf"])].copy()
    merged = merged.sort_values("成交额", ascending=False).head(int(cfg["data"]["etf_candidate_size"])).reset_index(drop=True)
    return merged[["代码", "名称", "asset_type", "最新价", "成交额", "成交量"]].copy()


def combine_with_holdings(candidates: pd.DataFrame, holdings: pd.DataFrame) -> pd.DataFrame:
    combined = candidates.copy()
    if combined.empty and holdings.empty:
        return combined
    if not combined.empty:
        combined["asset_type"] = combined["asset_type"].map(lambda x: infer_asset_type("", x) or x)
    if holdings.empty:
        combined["shares"] = np.nan
        combined["cost_price"] = np.nan
        combined["is_holding"] = False
        return combined
    h = holdings.copy()
    h["is_holding"] = True
    h["spot_latest_price"] = np.nan
    h["spot_amount"] = np.nan
    h["spot_volume"] = np.nan
    h["latest_price_from_spot"] = np.nan
    h["from_holding_only"] = True
    h = h.rename(columns={"code": "代码", "name": "名称", "asset_type": "asset_type", "shares": "shares", "cost_price": "cost_price"})
    h["名称"] = h["名称"].map(clean_name)
    h["asset_type"] = h["asset_type"].map(lambda x: infer_asset_type("", x))
    h["最新价"] = np.nan
    h["成交额"] = np.nan
    h["成交量"] = np.nan
    h["from_holding_only"] = True
    if not combined.empty:
        combined = combined.merge(
            holdings[["asset_type", "code", "shares", "cost_price"]].rename(columns={"code": "代码"}),
            on=["asset_type", "代码"],
            how="left",
        )
        combined["is_holding"] = combined["shares"].notna()
        combined["from_holding_only"] = False
    else:
        combined["shares"] = np.nan
        combined["cost_price"] = np.nan
        combined["is_holding"] = False
        combined["from_holding_only"] = False
    existing_keys = set(zip(combined["asset_type"].astype(str), combined["代码"].astype(str)))
    extra_rows = []
    for _, row in h.iterrows():
        key = (str(row["asset_type"]), str(row["代码"]))
        if key in existing_keys:
            continue
        extra_rows.append(row.to_dict())
    if extra_rows:
        extra = pd.DataFrame(extra_rows)
        extra["is_holding"] = True
        extra["from_holding_only"] = True
        combined = pd.concat([combined, extra], ignore_index=True, sort=False)
    for col in ["最新价", "成交额", "成交量", "shares", "cost_price"]:
        if col not in combined.columns:
            combined[col] = np.nan
    combined["名称"] = combined["名称"].map(clean_name)
    combined = combined.drop_duplicates(subset=["asset_type", "代码"], keep="first").reset_index(drop=True)
    return combined


def yahoo_symbol_for_stock(code: str) -> str:
    code = normalize_code(code)
    if code.startswith("6"):
        return f"{code}.SS"
    return f"{code}.SZ"


def fetch_yahoo_history(asset_type: str, code: str, cfg: Dict[str, Any]) -> pd.DataFrame:
    if asset_type != "STOCK":
        raise ValueError("Yahoo fallback currently supports STOCK only")
    days = max(int(cfg["data"].get("history_days", 300)), int(cfg["filter"].get("min_history_bars", 120)) * 2)
    range_text = "2y" if days > 370 else "1y"
    symbol = yahoo_symbol_for_stock(code)
    url = (
        f"https://query1.finance.yahoo.com/v8/finance/chart/{urllib.parse.quote(symbol)}"
        f"?range={range_text}&interval=1d&events=history|div|split&includeAdjustedClose=true"
    )
    req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
    with urllib.request.urlopen(req, timeout=30) as resp:
        payload = json.loads(resp.read().decode("utf-8"))
    chart = payload.get("chart") or {}
    if chart.get("error"):
        raise RuntimeError(chart["error"])
    result = (chart.get("result") or [None])[0]
    if not result:
        raise RuntimeError("Yahoo历史行情返回为空")
    timestamps = result.get("timestamp") or []
    quote = ((result.get("indicators") or {}).get("quote") or [{}])[0]
    if not timestamps or not quote:
        raise RuntimeError("Yahoo历史行情缺少OHLCV字段")
    out = pd.DataFrame({
        "date": pd.to_datetime(timestamps, unit="s", errors="coerce").date,
        "open": quote.get("open"),
        "close": quote.get("close"),
        "high": quote.get("high"),
        "low": quote.get("low"),
        "amount": quote.get("volume"),
    })
    for col in ["open", "close", "high", "low", "amount"]:
        out[col] = pd.to_numeric(out[col], errors="coerce")
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date", "close", "amount"]).sort_values("date").reset_index(drop=True)
    out = out.tail(int(cfg["data"].get("history_days", 300))).copy()
    out["turnover_amount"] = out["amount"] * out["close"] * 100.0
    return out[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()


def build_history_frame(asset_type: str, code: str, cfg: Dict[str, Any]) -> pd.DataFrame:
    cached = read_history_cache(asset_type, code, cfg)
    if cached is not None:
        return cached
    today = current_date_obj()
    start = (today - timedelta(days=int(cfg["data"]["history_days"]))).strftime("%Y%m%d")
    end = today.strftime("%Y%m%d")
    if asset_type == "STOCK":
        if str(cfg["data"].get("stock_history_source", "yahoo")).lower() == "yahoo":
            out = fetch_yahoo_history(asset_type, code, cfg)
            write_history_cache(asset_type, code, cfg, out)
            return out
        symbol = symbol_for_stock(code)
        try:
            df = call_with_retry(
                ak.stock_zh_a_hist_tx,
                symbol=symbol,
                start_date=start,
                end_date=end,
                adjust=str(cfg["data"]["adjust"]),
                retries=1,
            )
        except Exception as first_exc:
            try:
                df = call_with_retry(
                    ak.stock_zh_a_hist,
                    symbol=code,
                    period="daily",
                    start_date=start,
                    end_date=end,
                    adjust=str(cfg["data"]["adjust"]),
                    retries=1,
                )
            except Exception as second_exc:
                try:
                    out = fetch_yahoo_history(asset_type, code, cfg)
                    write_history_cache(asset_type, code, cfg, out)
                    return out
                except Exception as yahoo_exc:
                    raise RuntimeError(
                        f"历史行情接口均失败：tx={first_exc}; ak={second_exc}; yahoo={yahoo_exc}"
                    )
        df = df.copy()
        if "date" not in df.columns:
            raise RuntimeError("股票历史数据缺少日期字段")
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        for col in ["open", "close", "high", "low", "amount"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        df = df.dropna(subset=["date", "close", "amount"]).sort_values("date").reset_index(drop=True)
        df["turnover_amount"] = df["amount"] * df["close"] * 100.0
        out = df[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()
        write_history_cache(asset_type, code, cfg, out)
        return out
    if asset_type == "ETF":
        symbol = symbol_for_etf(code)
        try:
            df = call_with_retry(ak.fund_etf_hist_sina, symbol=symbol, retries=1)
        except Exception:
            df = call_with_retry(
                ak.fund_etf_hist_em,
                symbol=code,
                period="daily",
                start_date="19700101",
                end_date=end,
                adjust=str(cfg["data"]["adjust"]),
                retries=1,
            )
        df = df.copy()
        if "date" not in df.columns:
            raise RuntimeError("ETF历史数据缺少日期字段")
        df["date"] = pd.to_datetime(df["date"], errors="coerce")
        for col in ["open", "close", "high", "low", "amount"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")
        df = df.dropna(subset=["date", "close", "amount"]).sort_values("date").reset_index(drop=True)
        df["turnover_amount"] = df["amount"]
        out = df[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()
        write_history_cache(asset_type, code, cfg, out)
        return out
    raise ValueError(f"不支持的标的类型：{asset_type}")


def compute_indicators(hist: pd.DataFrame) -> Dict[str, Any]:
    close = hist["close"].astype(float).reset_index(drop=True)
    turnover = hist["turnover_amount"].astype(float).reset_index(drop=True)
    latest_close = float(close.iloc[-1])
    ma20 = float(close.rolling(20).mean().iloc[-1])
    ma60 = float(close.rolling(60).mean().iloc[-1])
    ma120 = float(close.rolling(120).mean().iloc[-1])
    ret20 = float(close.iloc[-1] / close.iloc[-20] - 1.0)
    ret60 = float(close.iloc[-1] / close.iloc[-60] - 1.0)
    ret120 = float(close.iloc[-1] / close.iloc[-120] - 1.0)
    vol60 = float(close.pct_change().tail(60).std(ddof=0) * math.sqrt(252))
    last60 = close.tail(60).reset_index(drop=True)
    drawdown = last60 / last60.cummax() - 1.0
    mdd60 = float(abs(drawdown.min()))
    avg_amount_20 = float(turnover.tail(20).mean())
    avg_amount_60 = float(turnover.tail(60).mean())
    amount_ratio = float(avg_amount_20 / avg_amount_60) if avg_amount_60 > 0 else np.nan
    above_ma60 = bool(latest_close >= ma60)
    above_ma120 = bool(latest_close >= ma120)
    ma20_above_ma60 = bool(ma20 >= ma60)
    price_ma60_ratio = float(latest_close / ma60)
    ma20_ma60_ratio = float(ma20 / ma60)
    price_ma120_ratio = float(latest_close / ma120)
    trend_ok = above_ma60 and ma20_above_ma60
    return {
        "latest_close": latest_close,
        "ma20": ma20,
        "ma60": ma60,
        "ma120": ma120,
        "ret20": ret20,
        "ret60": ret60,
        "ret120": ret120,
        "vol60": vol60,
        "mdd60": mdd60,
        "avg_amount_20": avg_amount_20,
        "avg_amount_60": avg_amount_60,
        "amount_ratio": amount_ratio,
        "above_ma60": above_ma60,
        "above_ma120": above_ma120,
        "ma20_above_ma60": ma20_above_ma60,
        "price_ma60_ratio": price_ma60_ratio,
        "ma20_ma60_ratio": ma20_ma60_ratio,
        "price_ma120_ratio": price_ma120_ratio,
        "trend_ok": trend_ok,
        "hist_bars": int(len(hist)),
    }


def apply_spot_price_to_history(hist: pd.DataFrame, rec: Dict[str, Any], cfg: Dict[str, Any]) -> pd.DataFrame:
    """Ensure weekday runs use the current spot price as the latest bar.

    The historical K-line cache is useful for moving averages, but a trading-day
    stock selection run must not rank candidates on an old close. If today's
    spot snapshot is available, append/replace today's last bar with that spot
    price. If it is a weekday and neither today's history nor spot price is
    available, fail fast instead of silently using stale cached prices.
    """
    out = hist.copy()
    if out.empty:
        return out
    today = current_date_obj().date()
    is_weekday = current_date_obj().weekday() < 5
    strict_fresh = bool(cfg.get("data", {}).get("require_fresh_price_on_weekday", True))
    out["date"] = pd.to_datetime(out["date"], errors="coerce")
    out = out.dropna(subset=["date", "close"]).sort_values("date").reset_index(drop=True)
    last_date = out["date"].iloc[-1].date()
    spot_price = to_number(rec.get("spot_latest_price"))
    spot_amount = to_number(rec.get("spot_amount"))
    spot_volume = to_number(rec.get("spot_volume"))
    has_spot = pd.notna(spot_price) and float(spot_price) > 0
    if has_spot:
        new_bar = {
            "date": pd.to_datetime(today),
            "open": float(spot_price),
            "close": float(spot_price),
            "high": float(spot_price),
            "low": float(spot_price),
            "amount": float(spot_volume) if pd.notna(spot_volume) and float(spot_volume) > 0 else float(out["amount"].iloc[-1]),
            "turnover_amount": float(spot_amount) if pd.notna(spot_amount) and float(spot_amount) > 0 else float(spot_price) * float(out["amount"].iloc[-1]) * 100.0,
        }
        if last_date == today:
            for col, value in new_bar.items():
                out.loc[out.index[-1], col] = value
        elif is_weekday:
            out = pd.concat([out, pd.DataFrame([new_bar])], ignore_index=True, sort=False)
        out = out.sort_values("date").reset_index(drop=True)
        return out[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()
    if is_weekday and strict_fresh and last_date != today:
        raise RuntimeError(f"历史行情不是最新交易日：last={last_date}, today={today}，且无现货最新价可覆盖")
    return out[["date", "open", "close", "high", "low", "amount", "turnover_amount"]].copy()


def evaluate_base_filters(row: Dict[str, Any], cfg: Dict[str, Any]) -> List[str]:
    reasons: List[str] = []
    asset_type = str(row.get("asset_type", row.get("标的类型", ""))).upper()
    code = normalize_code(row.get("代码") or row.get("code"))
    name = row.get("名称", row.get("name", ""))
    latest_close = to_number(row.get("latest_close"))
    avg_amount_20 = to_number(row.get("avg_amount_20"))
    ret20 = to_number(row.get("ret20"))
    mdd60 = to_number(row.get("mdd60"))
    ma60 = to_number(row.get("ma60"))
    ma20 = to_number(row.get("ma20"))
    hist_bars = int(row.get("hist_bars") or 0)
    min_price = float(cfg["filter"]["min_price_stock"]) if asset_type == "STOCK" else float(cfg["filter"]["min_price_etf"])
    min_amount = float(cfg.get("universe", {}).get("min_avg_amount_20d", cfg["filter"]["min_amount_stock"])) if asset_type == "STOCK" else float(cfg["filter"]["min_amount_etf"])
    if not ((is_allowed_stock(code) and asset_type == "STOCK") or (is_allowed_etf(code) and asset_type == "ETF")):
        reasons.append("代码不符合投资范围")
    if is_st_like(name):
        reasons.append("名称包含ST")
    if hist_bars < int(cfg["filter"]["min_history_bars"]):
        reasons.append("历史数据不足120个交易日")
    if pd.notna(latest_close) and latest_close < min_price:
        reasons.append("最新价低于最低价格要求")
    if pd.notna(avg_amount_20) and avg_amount_20 < min_amount:
        reasons.append("近20日日均成交额不足")
    if pd.notna(ret20) and ret20 > float(cfg["filter"]["max_ret20"]):
        reasons.append("近20日涨幅超过45%")
    if pd.notna(mdd60) and mdd60 > float(cfg["filter"]["max_mdd60_abs"]):
        reasons.append("近60日最大回撤超过35%")
    if pd.notna(latest_close) and pd.notna(ma60) and pd.notna(ma20):
        if latest_close < ma60 and ma20 < ma60:
            reasons.append("价格跌破60日均线且20日均线低于60日均线")
    if asset_type == "STOCK":
        qcfg = cfg.get("quality", {}).get("hard_filter", {})
        roe = to_number(row.get("ROE"))
        ocf = to_number(row.get("经营现金流/净利润"))
        growth = to_number(row.get("扣非净利润同比"))
        if pd.notna(roe) and roe < float(qcfg.get("roe_min", 0)):
            reasons.append("ROE为负")
        if pd.notna(ocf) and ocf < float(qcfg.get("ocf_to_net_profit_min", 0)):
            reasons.append("经营现金流/净利润为负")
        if pd.notna(growth) and growth < float(qcfg.get("deducted_profit_growth_min", -0.50)):
            reasons.append("扣非净利润同比低于-50%")
    return reasons


def pct_rank(series: pd.Series, higher_is_better: bool = True) -> pd.Series:
    clean = pd.to_numeric(series, errors="coerce")
    if higher_is_better:
        ranked = clean.rank(pct=True, method="average", ascending=True)
    else:
        ranked = (-clean).rank(pct=True, method="average", ascending=True)
    return ranked.fillna(0) * 100.0


def score_candidates(df: pd.DataFrame, cfg: Dict[str, Any]) -> pd.DataFrame:
    if df.empty:
        return df.copy()
    out = df.copy()
    for col in ["ROE", "经营现金流/净利润", "扣非净利润同比"]:
        if col not in out.columns:
            out[col] = np.nan
        out[col] = pd.to_numeric(out[col], errors="coerce")
    out["pct_ret20"] = pct_rank(out["ret20"], higher_is_better=True)
    out["pct_ret60"] = pct_rank(out["ret60"], higher_is_better=True)
    out["pct_ret120"] = pct_rank(out["ret120"], higher_is_better=True)
    out["pct_price_ma60_ratio"] = pct_rank(out["price_ma60_ratio"], higher_is_better=True)
    out["pct_ma20_ma60_ratio"] = pct_rank(out["ma20_ma60_ratio"], higher_is_better=True)
    out["pct_price_ma120_ratio"] = pct_rank(out["price_ma120_ratio"], higher_is_better=True)
    out["pct_amount_ratio"] = pct_rank(out["amount_ratio"], higher_is_better=True)
    out["pct_avg_amount_20"] = pct_rank(out["avg_amount_20"], higher_is_better=True)
    out["pct_vol60"] = pct_rank(out["vol60"], higher_is_better=False)
    out["pct_mdd60"] = pct_rank(out["mdd60"], higher_is_better=False)

    weights = cfg.get("factor_weights", {})
    out["score_momentum"] = (
        out["pct_ret20"] * 0.30
        + out["pct_ret60"] * 0.45
        + out["pct_ret120"] * 0.25
    )
    out["score_trend"] = (
        out["pct_price_ma60_ratio"] * 0.40
        + out["pct_ma20_ma60_ratio"] * 0.30
        + out["pct_price_ma120_ratio"] * 0.30
    )
    out["score_momentum_trend"] = out["score_momentum"] * 0.55 + out["score_trend"] * 0.45
    out["score_volume"] = out["pct_amount_ratio"]
    out["score_liquidity"] = out["pct_avg_amount_20"] * 0.70 + out["pct_amount_ratio"] * 0.30
    out["score_risk"] = out["pct_vol60"] * 0.50 + out["pct_mdd60"] * 0.50
    out["valuation_score"] = compute_valuation_score(out, cfg)
    out["total_score"] = (
        out["score_momentum_trend"] * float(weights.get("momentum_trend", weights.get("momentum", 0.4375)))
        + out["valuation_score"] * float(weights.get("valuation", 0.1875))
        + out["score_liquidity"] * float(weights.get("liquidity", 0.1875))
        + out["score_risk"] * float(weights.get("risk", 0.1875))
    )
    out["入选原因"] = out.apply(
        lambda r: f"趋势动量{r.get('score_momentum_trend', 0):.1f}，估值{r.get('valuation_score', 0):.1f}，流动性{r.get('score_liquidity', 0):.1f}，风险{r.get('score_risk', 0):.1f}",
        axis=1,
    )
    out["风险提示"] = out.apply(
        lambda r: "；".join([x for x in [
            "PE缺失或非正" if pd.isna(r.get("pe_ttm")) or to_number(r.get("pe_ttm")) <= 0 else "",
            "PE偏高" if to_number(r.get("pe_ttm")) > 80 else "",
            "上市日期缺失" if pd.isna(r.get("listing_days")) else "",
        ] if x]) or "无明显模型风险提示",
        axis=1,
    )
    out = out.sort_values(["total_score", "avg_amount_20"], ascending=[False, False]).reset_index(drop=True)
    out["rank"] = np.arange(1, len(out) + 1)
    return out


def calc_market_vitality(df: pd.DataFrame, cfg: Dict[str, Any]) -> Dict[str, Any]:
    if df.empty:
        return {
            "score": 0,
            "state": "无有效候选",
            "target_position": float(cfg["strategy"]["target_position_weak"]),
            "signals": {},
        }
    above_ma60_ratio = float(df["above_ma60"].map(truthy).mean())
    above_ma120_ratio = float(df["above_ma120"].map(truthy).mean())
    positive_ret20_ratio = float((df["ret20"] > 0).mean())
    positive_ret60_ratio = float((df["ret60"] > 0).mean())
    amount_boost_ratio = float((df["amount_ratio"] >= 1).mean())
    score = 0
    score += 1 if above_ma60_ratio >= 0.50 else 0
    score += 1 if above_ma120_ratio >= 0.45 else 0
    score += 1 if positive_ret20_ratio >= 0.55 else 0
    score += 1 if positive_ret60_ratio >= 0.50 else 0
    score += 1 if amount_boost_ratio >= 0.50 else 0
    if score <= 1:
        state = "弱市场"
        target_position = float(cfg["strategy"]["target_position_weak"])
    elif score <= 3:
        state = "普通市场"
        target_position = float(cfg["strategy"]["target_position_normal"])
    elif score == 4:
        state = "强市场"
        target_position = float(cfg["strategy"]["target_position_strong"])
    else:
        state = "活跃市场"
        target_position = float(cfg["strategy"]["target_position_hot"])
    return {
        "score": score,
        "state": state,
        "target_position": target_position,
        "signals": {
            "站上60日线比例": above_ma60_ratio,
            "站上120日线比例": above_ma120_ratio,
            "近20日收益为正比例": positive_ret20_ratio,
            "近60日收益为正比例": positive_ret60_ratio,
            "成交额放大比例": amount_boost_ratio,
            "参与判断标的数": int(len(df)),
        },
    }


def format_pct(value: Optional[float], digits: int = 1) -> str:
    if value is None or pd.isna(value):
        return "-"
    return f"{float(value):.{digits}%}"


def round_down_to_lot(value: Any, lot_size: int = 100) -> int:
    num = to_number(value)
    if pd.isna(num) or num <= 0 or lot_size <= 0:
        return 0
    return int(math.floor(num / lot_size) * lot_size)


def estimate_buy_shares(latest_close: Any, target_weight: Any, reference_capital: float, lot_size: int = 100) -> int:
    px = to_number(latest_close)
    weight = to_number(target_weight)
    if pd.isna(px) or px <= 0 or pd.isna(weight) or weight <= 0 or reference_capital <= 0:
        return 0
    target_value = float(reference_capital) * float(weight)
    return round_down_to_lot(target_value / px, lot_size)


def estimate_reduce_shares(current_shares: Any, lot_size: int = 100) -> int:
    shares = to_share_int(current_shares)
    if shares <= 0:
        return 0
    if shares < lot_size:
        return shares
    half_lot = round_down_to_lot(shares * 0.5, lot_size)
    return min(shares, max(lot_size, half_lot))


def build_trade_instruction(buy_shares: int = 0, sell_shares: int = 0) -> str:
    if buy_shares > 0 and sell_shares > 0:
        return f"买入 {buy_shares} 股，卖出 {sell_shares} 股"
    if buy_shares > 0:
        return f"买入 {buy_shares} 股"
    if sell_shares > 0:
        return f"卖出 {sell_shares} 股"
    return "不操作"


def recommend_candidate_trade(row: pd.Series, reference_capital: float, lot_size: int = 100) -> Tuple[int, float, str]:
    action = str(row.get("action", ""))
    latest_close = to_number(row.get("latest_close"))
    target_weight = to_number(row.get("target_weight"))
    if action != "可新开仓":
        return 0, np.nan, "不操作"
    buy_shares = estimate_buy_shares(latest_close, target_weight, reference_capital, lot_size)
    if buy_shares <= 0:
        return 0, np.nan, "资金不足一手，暂不买入"
    trade_value = float(buy_shares) * float(latest_close)
    return buy_shares, trade_value, build_trade_instruction(buy_shares=buy_shares)


def recommend_holding_trade(row: pd.Series, lot_size: int = 100) -> Tuple[int, float, str]:
    action = str(row.get("action", ""))
    latest_close = to_number(row.get("latest_close"))
    current_shares = to_share_int(row.get("shares"))
    if action == "建议卖出":
        sell_shares = current_shares
    elif action == "建议减仓":
        sell_shares = estimate_reduce_shares(current_shares, lot_size)
    else:
        sell_shares = 0
    if sell_shares <= 0:
        return 0, np.nan, "不操作"
    trade_value = float(sell_shares) * float(latest_close) if pd.notna(latest_close) else np.nan
    return sell_shares, trade_value, build_trade_instruction(sell_shares=sell_shares)


def recommend_reduce_trade(row: pd.Series, lot_size: int = 100) -> Tuple[int, float, str]:
    latest_close = to_number(row.get("latest_close"))
    sell_shares = estimate_reduce_shares(row.get("shares"), lot_size)
    if sell_shares <= 0:
        return 0, np.nan, "不操作"
    trade_value = float(sell_shares) * float(latest_close) if pd.notna(latest_close) else np.nan
    return sell_shares, trade_value, build_trade_instruction(sell_shares=sell_shares)


def recommend_sell_all_trade(row: pd.Series) -> Tuple[int, float, str]:
    latest_close = to_number(row.get("latest_close"))
    current_shares = to_share_int(row.get("shares"))
    if current_shares <= 0:
        return 0, np.nan, "不操作"
    trade_value = float(current_shares) * float(latest_close) if pd.notna(latest_close) else np.nan
    return current_shares, trade_value, build_trade_instruction(sell_shares=current_shares)


def current_holding_market_value(holdings_check: pd.DataFrame) -> float:
    if holdings_check.empty:
        return 0.0
    latest = pd.to_numeric(holdings_check.get("latest_close"), errors="coerce")
    cost = pd.to_numeric(holdings_check.get("cost_price"), errors="coerce")
    shares = pd.to_numeric(holdings_check.get("shares"), errors="coerce").fillna(0.0)
    price = latest.where(latest.notna() & (latest > 0), cost)
    value = (shares * price.fillna(0.0)).sum()
    return float(value) if pd.notna(value) else 0.0


def cash_balance(holdings: pd.DataFrame) -> float:
    if holdings.empty or "asset_type" not in holdings.columns:
        return 0.0
    cash_rows = holdings[holdings["asset_type"].astype(str).str.upper() == "CASH"].copy()
    if cash_rows.empty:
        return 0.0
    shares = pd.to_numeric(cash_rows.get("shares"), errors="coerce").fillna(0.0)
    cost = pd.to_numeric(cash_rows.get("cost_price"), errors="coerce").fillna(1.0)
    value = (shares * cost).sum()
    return float(value) if pd.notna(value) else 0.0


def build_account_summary(holdings: pd.DataFrame, holdings_check: pd.DataFrame, fallback_capital: float) -> Dict[str, float]:
    equity_value = current_holding_market_value(holdings_check)
    cash = cash_balance(holdings)
    total_asset = equity_value + cash
    unrealized_pnl = 0.0
    holding_cost = 0.0
    if not holdings_check.empty:
        shares = pd.to_numeric(holdings_check.get("shares"), errors="coerce").fillna(0.0)
        cost = pd.to_numeric(holdings_check.get("cost_price"), errors="coerce").fillna(0.0)
        latest = pd.to_numeric(holdings_check.get("latest_close"), errors="coerce")
        holding_cost = float((shares * cost).sum())
        pnl_amount = (latest - cost) * shares
        unrealized_pnl = float(pnl_amount.dropna().sum()) if not pnl_amount.dropna().empty else 0.0
    if total_asset <= 0:
        total_asset = float(fallback_capital)
        cash = max(0.0, total_asset - equity_value)
    current_position = equity_value / total_asset if total_asset > 0 else 0.0
    unrealized_pnl_pct = unrealized_pnl / holding_cost if holding_cost > 0 else 0.0
    return {
        "current_equity_value": float(equity_value),
        "cash_balance": float(cash),
        "total_asset": float(total_asset),
        "current_position": float(current_position),
        "holding_cost": float(holding_cost),
        "unrealized_pnl": float(unrealized_pnl),
        "unrealized_pnl_pct": float(unrealized_pnl_pct),
    }


def build_actual_buy_plan(
    scored: pd.DataFrame,
    total_asset: float,
    target_position: float,
    top_n: int,
    existing_equity_value: float = 0.0,
    available_cash: float = 0.0,
    lot_size: int = 100,
) -> Tuple[pd.DataFrame, Dict[str, Dict[str, Any]]]:
    target_equity_value = float(total_asset) * float(target_position)
    position_gap = max(0.0, target_equity_value - float(existing_equity_value))
    budget = min(position_gap, max(0.0, float(available_cash)))
    if scored.empty or budget <= 0:
        return pd.DataFrame(), {}
    candidates = scored.copy()
    candidates = candidates[candidates["action"] == "可新开仓"].copy()
    if candidates.empty:
        return pd.DataFrame(), {}
    candidates["latest_close_num"] = pd.to_numeric(candidates["latest_close"], errors="coerce")
    candidates["lot_cost"] = candidates["latest_close_num"] * float(lot_size)
    candidates = candidates[
        candidates["latest_close_num"].notna()
        & (candidates["latest_close_num"] > 0)
        & candidates["lot_cost"].notna()
        & (candidates["lot_cost"] <= budget)
    ].copy()
    if candidates.empty:
        return pd.DataFrame(), {}
    candidates = candidates.sort_values(["rank", "total_score", "latest_close_num"], ascending=[True, False, True]).reset_index(drop=True)
    selected = candidates.head(int(max(1, top_n))).copy()
    if selected.empty:
        return pd.DataFrame(), {}

    remaining_budget = budget
    records = selected.to_dict("records")
    for rec in records:
        rec["buy_shares"] = 0

    while True:
        bought_any = False
        for rec in records:
            lot_cost = float(rec["lot_cost"])
            if remaining_budget >= lot_cost:
                rec["buy_shares"] = int(rec["buy_shares"]) + int(lot_size)
                remaining_budget -= lot_cost
                bought_any = True
        if not bought_any:
            break

    plan_rows: List[Dict[str, Any]] = []
    lookup: Dict[str, Dict[str, Any]] = {}
    for rec in records:
        buy_shares = int(rec.get("buy_shares", 0))
        if buy_shares <= 0:
            continue
        latest_close = float(rec["latest_close_num"])
        trade_value = float(buy_shares) * latest_close
        code = str(rec.get("code", ""))
        row = {
            "rank": rec.get("rank"),
            "code": code,
            "name": rec.get("name", ""),
            "asset_type": rec.get("asset_type", ""),
            "is_holding": rec.get("is_holding", ""),
            "latest_close": latest_close,
            "total_score": rec.get("total_score", np.nan),
            "score_momentum": rec.get("score_momentum", np.nan),
            "score_trend": rec.get("score_trend", np.nan),
            "score_volume": rec.get("score_volume", np.nan),
            "score_liquidity": rec.get("score_liquidity", np.nan),
            "score_risk": rec.get("score_risk", np.nan),
            "target_weight": rec.get("target_weight", np.nan),
            "buy_shares": buy_shares,
            "trade_value": trade_value,
            "trade_instruction": build_trade_instruction(buy_shares=buy_shares),
            "action": "可新开仓",
            "reason": rec.get("reason", ""),
        }
        plan_rows.append(row)
        lookup[code] = {
            "buy_shares": buy_shares,
            "trade_value": trade_value,
            "trade_instruction": row["trade_instruction"],
        }
    plan_df = pd.DataFrame(plan_rows)
    if not plan_df.empty:
        plan_df = plan_df.sort_values(["rank", "total_score"], ascending=[True, False]).reset_index(drop=True)
    return plan_df, lookup


def build_core_recommendation(
    market: Dict[str, Any],
    buy_count: int,
    reduce_count: int,
    sell_count: int,
    top_n: int,
    total_asset: float,
) -> str:
    state = market["state"]
    position = market["target_position"]
    per_position = position / float(max(1, top_n))
    if state == "弱市场":
        base = f"市场偏弱，建议权益仓位控制在{position:.0%}，优先处理风险较高持仓，谨慎新增仓位。"
    elif state == "普通市场":
        base = f"市场正常，建议权益仓位约{position:.0%}，以高分候选和现有核心持仓为主，稳步调整。"
    elif state == "强市场":
        base = f"市场较强，建议权益仓位约{position:.0%}，可以围绕前10名高分标的做更积极的配置。"
    elif state == "活跃市场":
        base = f"市场活跃，建议权益仓位约{position:.0%}，优先关注趋势和动量都较好的候选标的。"
    else:
        base = "当前没有足够有效候选，先检查数据源和持仓文件。"
    instruction_lines = [f"账户总资产：{total_asset:,.0f} 元。", "操作顺序：先卖出，再减仓，最后买入。"]
    instruction_lines.append(f"建议卖出：{sell_count}只。")
    instruction_lines.append(f"建议减仓：{reduce_count}只。")
    instruction_lines.append(f"可新开仓：{buy_count}只。")
    instruction_lines.append(f"单只目标仓位：{per_position:.0%}。")
    instruction_lines.append("买入会优先选择预算内可以实际成交的最佳标的，单价过高的标的会自动跳过。")
    instruction_lines.append("减仓股数按当前持仓的一半估算。")
    return base + "\n" + "\n".join(instruction_lines)


def decide_candidate_action(row: pd.Series, top_n: int) -> Tuple[str, str]:
    if truthy(row.get("is_holding", False)):
        if int(row.get("rank", 10**9)) <= top_n:
            return "继续持有", f"综合评分进入前{top_n}，且当前已在持仓中，可以继续持有。"
        return "观察", f"该标的当前已持有，但综合排名未进入前{top_n}，先观察后续变化。"
    return "可新开仓", f"综合评分进入前{top_n}，且不在当前持仓中，可以作为新开仓候选。"


def decide_hold_action(row: pd.Series, cfg: Dict[str, Any]) -> Tuple[str, str]:
    rank = row.get("rank")
    pnl = row.get("pnl_pct")
    trend_ok = bool(row.get("trend_ok", False))
    base_failed = bool(row.get("base_failed", False))
    latest_close = to_number(row.get("latest_close"))
    ma60 = to_number(row.get("ma60"))
    if base_failed:
        return "建议卖出", "触发基础过滤，标的风险过高或不符合投资范围。"
    if pd.notna(pnl) and pnl <= float(cfg["risk"]["stop_loss_pct"]):
        return "建议卖出", "浮动亏损超过15%，需要优先控制回撤。"
    if rank is None or pd.isna(rank):
        return "建议卖出", "无法获得有效排名，建议先降低不确定性。"
    sell_buffer = int(cfg.get("rebalance", {}).get("sell_rank_buffer", cfg.get("strategy", {}).get("sell_rank_buffer", 30)))
    hard_sell_rank = max(sell_buffer + 20, 50)
    if int(rank) > hard_sell_rank:
        return "建议卖出", f"综合评分跌出前{hard_sell_rank}，后续竞争力不足。"
    if pd.notna(pnl) and pnl <= float(cfg["risk"]["warning_loss_pct"]):
        return "建议减仓", "浮动亏损超过10%，建议先降低仓位观察。"
    if pd.notna(latest_close) and pd.notna(ma60) and latest_close < ma60:
        return "建议减仓", "价格已经跌破60日线，趋势结构转弱。"
    if int(rank) > sell_buffer:
        return "建议减仓", f"综合评分跌出前{sell_buffer}，建议降低持仓权重。"
    buy_top_n = int(cfg.get("rebalance", {}).get("buy_top_n", cfg.get("strategy", {}).get("top_n", 10)))
    if int(rank) > buy_top_n:
        return "观察", f"综合评分仍在前{sell_buffer}内，但尚未进入前{buy_top_n}，先观察。"
    if trend_ok:
        return "继续持有", f"综合评分进入前{buy_top_n}，且价格位于60日均线上方，趋势结构较好。"
    return "建议减仓", "虽然排名较高，但趋势结构不够稳定，建议适度减仓。"


def chinese_asset_type(value: Any) -> str:
    text = str(value).upper()
    if text == "STOCK":
        return "股票"
    if text == "ETF":
        return "ETF"
    return str(value)


def chinese_bool(value: Any) -> str:
    if isinstance(value, str) and value in {"是", "否", "-"}:
        return value
    if pd.isna(value):
        return "-"
    return "是" if bool(value) else "否"


def truthy(value: Any) -> bool:
    if isinstance(value, str):
        return value.strip() in {"是", "True", "true", "1", "Y", "YES"}
    if pd.isna(value):
        return False
    return bool(value)


def clean_output_value(value: Any) -> Any:
    if value is None:
        return None
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (np.integer,)):
        return int(value)
    if isinstance(value, (np.floating,)):
        return float(value)
    if isinstance(value, (np.bool_, bool)):
        return bool(value)
    return value


def export_columns(df: pd.DataFrame, column_order: List[str]) -> pd.DataFrame:
    out = df.copy()
    for col in column_order:
        if col not in out.columns:
            out[col] = np.nan
    out = out[column_order].copy()
    rename_map = {
        "code": "代码",
        "name": "名称",
        "asset_type": "标的类型",
        "latest_close": "最新价",
        "ma20": "20日均线",
        "ma60": "60日均线",
        "ma120": "120日均线",
        "ret20": "近20日涨跌幅",
        "ret60": "近60日涨跌幅",
        "ret120": "近120日涨跌幅",
        "vol60": "近60日年化波动率",
        "mdd60": "近60日最大回撤",
        "avg_amount_20": "近20日日均成交额",
        "avg_amount_60": "近60日日均成交额",
        "amount_ratio": "成交额放大倍数",
        "above_ma60": "是否站上60日线",
        "above_ma120": "是否站上120日线",
        "ma20_above_ma60": "20日线是否高于60日线",
        "price_ma60_ratio": "最新价/60日均线",
        "ma20_ma60_ratio": "20日均线/60日均线",
        "price_ma120_ratio": "最新价/120日均线",
        "universe_score": "股票池排名分",
        "float_market_cap_score": "流通市值排名分",
        "amount_rank_score": "成交额排名分",
        "universe_rank": "股票池初筛排名",
        "股票池初筛排名": "股票池初筛排名",
        "float_market_cap": "流通市值",
        "listing_days": "上市天数",
        "industry": "行业",
        "ROE": "ROE",
        "经营现金流/净利润": "经营现金流/净利润",
        "扣非净利润同比": "扣非净利润同比",
        "pe_ttm": "PE_TTM",
        "peg": "PEG",
        "score_momentum_trend": "趋势动量分",
        "valuation_score": "估值得分",
        "入选原因": "入选原因",
        "风险提示": "风险提示",
        "score_momentum": "动量得分",
        "score_trend": "趋势得分",
        "score_volume": "成交活跃得分",
        "score_liquidity": "成交活跃分",
        "score_risk": "风险分",
        "total_score": "综合得分",
        "rank": "排名",
        "target_weight": "建议目标仓位",
        "action": "操作建议",
        "reason": "建议理由",
        "buy_shares": "建议买入股数",
        "sell_shares": "建议卖出股数",
        "trade_value": "建议交易金额",
        "trade_instruction": "建议执行",
        "shares": "持仓数量",
        "cost_price": "成本价",
        "pnl_amount": "浮动盈亏金额",
        "pnl_pct": "浮动盈亏率",
        "is_holding": "当前持仓",
        "reference_capital": "测试资金",
        "hist_bars": "历史交易日数",
        "spot_latest_price": "现货最新价",
        "spot_amount": "现货成交额",
        "spot_volume": "现货成交量",
        "stage": "失败阶段",
        "failure_reason": "失败原因",
    }
    out = out.rename(columns={k: v for k, v in rename_map.items() if k in out.columns})
    return out


def write_table_headers(ws, headers: List[str], row: int = 1) -> None:
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 24


def set_sheet_filter_and_freeze(ws, max_col: int, max_row: int, freeze_cell: str = "A2") -> None:
    ws.freeze_panes = freeze_cell
    if max_row >= 1 and max_col >= 1:
        ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"


def format_cell(cell, col_name: str) -> None:
    if col_name in {"近20日涨跌幅", "近60日涨跌幅", "近120日涨跌幅", "近60日年化波动率", "近60日最大回撤", "建议目标仓位", "浮动盈亏率", "ROE", "经营现金流/净利润", "扣非净利润同比"}:
        cell.number_format = "0.00%"
    elif col_name in {"最新价", "20日均线", "60日均线", "120日均线", "成本价", "成交价格"}:
        cell.number_format = "0.00"
    elif col_name in {"近20日日均成交额", "近60日日均成交额", "现货成交额", "浮动盈亏金额", "成交金额", "卖出匹配成本", "已实现盈亏", "流通市值"}:
        cell.number_format = '#,##0.00'
    elif col_name in {"成交额放大倍数", "最新价/60日均线", "20日均线/60日均线", "最新价/120日均线"}:
        cell.number_format = "0.00"
    elif col_name in {"动量得分", "趋势得分", "成交活跃得分", "流动性得分", "风险控制得分", "综合评分", "综合得分", "股票池排名分", "流通市值排名分", "成交额排名分", "估值得分", "趋势动量分", "成交活跃分", "风险分"}:
        cell.number_format = "0.0"
    elif col_name in {"排名", "历史交易日数", "持仓数量", "建议买入股数", "建议卖出股数"}:
        cell.number_format = "0"
    elif col_name in {"建议交易金额"}:
        cell.number_format = '#,##0.00'


def width_by_texts(values: Iterable[Any], min_width: int = 8, max_width: int = 48) -> int:
    longest = 0
    for value in values:
        if value is None:
            continue
        text = str(value)
        longest = max(longest, len(text))
    return max(min_width, min(max_width, int(longest * 1.15 + 2)))


def apply_row_fill(ws, row_idx: int, max_col: int, fill: Optional[PatternFill]) -> None:
    if fill is None:
        return
    for col_idx in range(1, max_col + 1):
        ws.cell(row=row_idx, column=col_idx).fill = fill


def write_dataframe_sheet(
    ws,
    df: pd.DataFrame,
    title: Optional[str] = None,
    action_col: Optional[str] = None,
    percent_cols: Optional[List[str]] = None,
    extra_widths: Optional[Dict[str, int]] = None,
    wrap_cols: Optional[List[str]] = None,
    start_row: int = 1,
    apply_filter: bool = True,
) -> None:
    percent_cols = percent_cols or []
    wrap_cols = wrap_cols or []
    extra_widths = extra_widths or {}
    if title:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=max(1, df.shape[1]))
        cell = ws.cell(row=start_row, column=1, value=title)
        cell.fill = TITLE_FILL
        cell.font = TITLE_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[start_row].height = 24
        header_row = start_row + 1
        data_start = start_row + 2
    else:
        header_row = start_row
        data_start = start_row + 1
    headers = list(df.columns)
    write_table_headers(ws, headers, row=header_row)
    for r_idx, (_, row) in enumerate(df.iterrows(), start=data_start):
        action = str(row.get(action_col, "")) if action_col else ""
        fill = PatternFill("solid", fgColor=ACTION_FILL_MAP.get(action, "FFFFFFFF")) if action in ACTION_FILL_MAP else None
        for c_idx, col_name in enumerate(headers, start=1):
            value = clean_output_value(row[col_name])
            if col_name in {"标的类型"}:
                value = chinese_asset_type(value)
            if col_name in {"是否站上60日线", "是否站上120日线", "20日线是否高于60日线", "当前持仓"}:
                value = chinese_bool(value)
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if col_name in wrap_cols or col_name in {"建议理由", "建议执行"}:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            elif col_name in {"名称", "建议理由"}:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center")
            format_cell(cell, col_name)
        apply_row_fill(ws, r_idx, len(headers), fill)
        if action in ACTION_FILL_MAP and action_col and action_col in headers:
            action_cell = ws.cell(row=r_idx, column=headers.index(action_col) + 1)
            action_cell.font = BODY_BOLD_FONT
    if apply_filter:
        freeze_cell = f"A{header_row + 1}"
        max_row = data_start + len(df) - 1
        set_sheet_filter_and_freeze(ws, len(headers), max_row, freeze_cell=freeze_cell)
    for c_idx, col_name in enumerate(headers, start=1):
        if col_name in extra_widths:
            width = extra_widths[col_name]
        else:
            values = [col_name] + [df.iloc[i, c_idx - 1] for i in range(min(len(df), 50))]
            if col_name in wrap_cols or col_name in {"建议理由"}:
                width = 42
            else:
                width = width_by_texts(values)
        ws.column_dimensions[get_column_letter(c_idx)].width = width


def add_summary_card(ws, start_row: int, start_col: int, title: str, value: str, subtitle: str = "", fill: str = "FFF8FAFC") -> None:
    end_row = start_row + 2
    end_col = start_col + 1
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=end_row, end_column=end_col)
    cell = ws.cell(row=start_row, column=start_col, value="\n".join([title, value, subtitle]).strip())
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.font = CARD_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=fill)
    ws.row_dimensions[start_row].height = 24
    ws.row_dimensions[start_row + 1].height = 22
    ws.row_dimensions[start_row + 2].height = 22


def write_key_value_table(ws, start_row: int, start_col: int, title: str, pairs: List[Tuple[str, str]], width: int = 30) -> int:
    ws.merge_cells(start_row=start_row, start_column=start_col, end_row=start_row, end_column=start_col + 3)
    title_cell = ws.cell(row=start_row, column=start_col, value=title)
    title_cell.fill = HEADER_FILL
    title_cell.font = HEADER_FONT
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    title_cell.border = THIN_BORDER
    row = start_row + 1
    for label, value in pairs:
        ws.cell(row=row, column=start_col, value=label).font = BODY_BOLD_FONT
        ws.cell(row=row, column=start_col).border = THIN_BORDER
        ws.cell(row=row, column=start_col).fill = PatternFill("solid", fgColor="FFF8FAFC")
        ws.cell(row=row, column=start_col + 1, value=value).font = BODY_FONT
        ws.cell(row=row, column=start_col + 1).border = THIN_BORDER
        ws.cell(row=row, column=start_col + 1).alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.cell(row=row, column=start_col + 1).fill = PatternFill("solid", fgColor="FFF8FAFC")
        for c in range(start_col + 2, start_col + 4):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFF8FAFC")
        row += 1
    for c in range(start_col, start_col + 4):
        ws.column_dimensions[get_column_letter(c)].width = width if c == start_col + 1 else 14
    return row


def write_action_section(ws, start_row: int, title: str, df: pd.DataFrame, fill_color: str) -> int:
    end_col = max(1, df.shape[1])
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=end_col)
    title_cell = ws.cell(row=start_row, column=1, value=title)
    title_cell.fill = PatternFill("solid", fgColor=fill_color)
    title_cell.font = Font(name="Microsoft YaHei", color="FF1F1F1F", bold=True, size=12)
    title_cell.alignment = Alignment(horizontal="left", vertical="center")
    title_cell.border = THIN_BORDER
    if df.empty:
        ws.merge_cells(start_row=start_row + 1, start_column=1, end_row=start_row + 1, end_column=end_col)
        empty_cell = ws.cell(row=start_row + 1, column=1, value="本节暂无数据。")
        empty_cell.alignment = Alignment(horizontal="left", vertical="center")
        empty_cell.font = BODY_FONT
        empty_cell.border = THIN_BORDER
        return start_row + 2
    write_dataframe_sheet(ws, df, title=None, action_col="操作建议", wrap_cols=["建议理由"])
    return start_row + len(df) + 2


def build_analysis_universe(cfg: Dict[str, Any], holdings: pd.DataFrame) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    failures: List[Dict[str, Any]] = []
    holdings = security_holdings(holdings)
    holdings = holdings[holdings["asset_type"].astype(str).str.upper() == "STOCK"].copy()
    stock_candidates = build_stock_candidates(cfg)
    if bool(cfg["data"].get("include_etf", False)):
        etf_candidates = build_etf_candidates(cfg)
    else:
        etf_candidates = pd.DataFrame(columns=["代码", "名称", "asset_type", "最新价", "成交额", "成交量", "总市值", "流通市值", "上市日期", "上市天数", "行业", "股票池排名分", "流通市值排名分", "成交额排名分", "PE_TTM", "PEG"])
    stock_candidates = stock_candidates.copy()
    etf_candidates = etf_candidates.copy()
    stock_candidates["from_candidate_pool"] = True
    etf_candidates["from_candidate_pool"] = True
    candidate_df = pd.concat([stock_candidates, etf_candidates], ignore_index=True, sort=False)
    candidate_df["asset_type"] = candidate_df["asset_type"].map(lambda x: infer_asset_type("", x) or x)
    candidate_df["shares"] = np.nan
    candidate_df["cost_price"] = np.nan
    candidate_df["is_holding"] = False
    candidate_df["from_holding_only"] = False
    candidate_df["spot_latest_price"] = pd.to_numeric(candidate_df["最新价"], errors="coerce")
    candidate_df["spot_amount"] = pd.to_numeric(candidate_df["成交额"], errors="coerce")
    candidate_df["spot_volume"] = pd.to_numeric(candidate_df["成交量"], errors="coerce")
    candidate_df["spot_market_cap"] = pd.to_numeric(candidate_df.get("总市值", np.nan), errors="coerce")
    candidate_df["pe_ttm"] = pd.to_numeric(candidate_df.get("PE_TTM", np.nan), errors="coerce")
    candidate_df["peg"] = pd.to_numeric(candidate_df.get("PEG", np.nan), errors="coerce")
    candidate_df["float_market_cap"] = pd.to_numeric(candidate_df.get("流通市值", np.nan), errors="coerce")
    candidate_df["listing_days"] = pd.to_numeric(candidate_df.get("上市天数", np.nan), errors="coerce")
    candidate_df["industry"] = candidate_df.get("行业", "未知")
    candidate_df["universe_score"] = pd.to_numeric(candidate_df.get("股票池排名分", np.nan), errors="coerce")
    candidate_df["float_market_cap_score"] = pd.to_numeric(candidate_df.get("流通市值排名分", np.nan), errors="coerce")
    candidate_df["amount_rank_score"] = pd.to_numeric(candidate_df.get("成交额排名分", np.nan), errors="coerce")
    candidate_df["universe_rank"] = pd.to_numeric(candidate_df.get("股票池初筛排名", np.nan), errors="coerce")
    for col in ["ROE", "经营现金流/净利润", "扣非净利润同比"]:
        candidate_df[col] = pd.to_numeric(candidate_df.get(col, np.nan), errors="coerce")

    if holdings.empty:
        return candidate_df, failures

    holdings = holdings.copy()
    holdings["key"] = holdings["asset_type"].astype(str) + "_" + holdings["code"].astype(str)
    candidate_df["key"] = candidate_df["asset_type"].astype(str) + "_" + candidate_df["代码"].astype(str)
    merged = candidate_df.merge(
        holdings[["key", "shares", "cost_price"]],
        on="key",
        how="left",
        suffixes=("", "_hold"),
    )
    merged["is_holding"] = merged["shares_hold"].notna()
    merged["shares"] = merged["shares_hold"]
    merged["cost_price"] = merged["cost_price_hold"]
    merged = merged.drop(columns=["shares_hold", "cost_price_hold"])
    candidate_df = merged

    existing_keys = set(candidate_df["key"].astype(str).tolist())
    extra_holding_rows = []
    for _, row in holdings.iterrows():
        key = str(row["key"])
        if key in existing_keys:
            continue
        extra_holding_rows.append(
            {
                "代码": row["code"],
                "名称": row["name"],
                "asset_type": row["asset_type"],
                "最新价": np.nan,
                "成交额": np.nan,
                "成交量": np.nan,
                "spot_latest_price": np.nan,
                "spot_amount": np.nan,
                "spot_volume": np.nan,
                "spot_market_cap": np.nan,
                "pe_ttm": np.nan,
                "peg": np.nan,
                "float_market_cap": np.nan,
                "listing_days": np.nan,
                "industry": "未知",
                "universe_score": np.nan,
                "float_market_cap_score": np.nan,
                "amount_rank_score": np.nan,
                "universe_rank": np.nan,
                "ROE": np.nan,
                "经营现金流/净利润": np.nan,
                "扣非净利润同比": np.nan,
                "shares": row["shares"],
                "cost_price": row["cost_price"],
                "is_holding": True,
                "from_holding_only": True,
                "from_candidate_pool": False,
                "key": key,
            }
        )
    if extra_holding_rows:
        candidate_df = pd.concat([candidate_df, pd.DataFrame(extra_holding_rows)], ignore_index=True, sort=False)
    candidate_df["名称"] = candidate_df["名称"].map(clean_name)
    candidate_df["asset_type"] = candidate_df["asset_type"].map(lambda x: infer_asset_type("", x) or x)
    candidate_df["代码"] = candidate_df["代码"].map(normalize_code)
    candidate_df = candidate_df.drop_duplicates(subset=["key"], keep="first").reset_index(drop=True)
    return candidate_df, failures


def early_filter_rows(df: pd.DataFrame, cfg: Dict[str, Any]) -> Tuple[pd.DataFrame, List[Dict[str, Any]]]:
    keep_rows = []
    failures: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        rec = row.to_dict()
        code = normalize_code(rec.get("代码"))
        asset_type = infer_asset_type(code, rec.get("asset_type"))
        name = clean_name(rec.get("名称"))
        rec["代码"] = code
        rec["asset_type"] = asset_type
        rec["名称"] = name
        rec["from_holding_only"] = bool(rec.get("from_holding_only", False))
        rec["is_holding"] = bool(rec.get("is_holding", False))
        spot_price = to_number(rec.get("spot_latest_price"))
        spot_amount = to_number(rec.get("spot_amount"))
        pe_ttm = to_number(rec.get("pe_ttm", rec.get("PE_TTM")))
        peg = to_number(rec.get("peg", rec.get("PEG")))
        if asset_type == "STOCK" and (pd.isna(pe_ttm) or pd.isna(peg)) and not bool(cfg["data"].get("disable_slow_fundamental_fetch", True)):
            try:
                if bool(cfg["data"].get("use_subprocess_fundamental_fetch", True)):
                    metrics = fetch_stock_valuation_metrics_subprocess(
                        code,
                        timeout_seconds=float(cfg["data"].get("fundamental_per_symbol_timeout_seconds", 8)),
                    )
                else:
                    metrics = fetch_stock_valuation_metrics(code)
                pe_ttm = to_number(metrics.get("PE_TTM"))
                peg = to_number(metrics.get("PEG"))
                rec["pe_ttm"] = pe_ttm
                rec["peg"] = peg
            except Exception:
                pe_ttm = np.nan
                peg = np.nan
        reasons: List[str] = []
        if asset_type == "STOCK" and not is_allowed_stock(code):
            reasons.append("代码不符合A股主板或创业板范围")
        if asset_type == "ETF" and not is_allowed_etf(code):
            reasons.append("代码不符合A股ETF范围")
        if is_st_like(name):
            reasons.append("名称包含ST")
        if pd.notna(spot_price):
            min_price = float(cfg["filter"]["min_price_stock"]) if asset_type == "STOCK" else float(cfg["filter"]["min_price_etf"])
            if spot_price < min_price:
                reasons.append("现货价格低于最低价格要求")
        if pd.notna(spot_amount):
            min_amount = float(cfg["filter"]["min_amount_stock"]) if asset_type == "STOCK" else float(cfg["filter"]["min_amount_etf"])
            if spot_amount < min_amount:
                reasons.append("现货成交额不足")
        # PE_TTM 仅用于估值得分，PEG 按当前配置禁用，不再作为硬过滤条件。
        if reasons:
            failures.append(
                {
                    "代码": code,
                    "名称": name,
                    "asset_type": asset_type,
                    "stage": "基础过滤",
                    "failure_reason": "；".join(reasons),
                }
            )
            continue
        keep_rows.append(rec)
    return pd.DataFrame(keep_rows), failures


def fetch_and_analyze_one(rec: Dict[str, Any], cfg: Dict[str, Any]) -> Tuple[Optional[Dict[str, Any]], Optional[Dict[str, Any]]]:
    code = normalize_code(rec.get("代码"))
    name = clean_name(rec.get("名称"))
    asset_type = infer_asset_type(code, rec.get("asset_type"))
    base = {
        "code": code,
        "name": name,
        "asset_type": asset_type,
        "is_holding": bool(rec.get("is_holding", False)),
        "shares": rec.get("shares"),
        "cost_price": rec.get("cost_price"),
        "spot_latest_price": rec.get("spot_latest_price"),
        "spot_amount": rec.get("spot_amount"),
        "spot_volume": rec.get("spot_volume"),
        "spot_market_cap": rec.get("spot_market_cap"),
        "pe_ttm": rec.get("pe_ttm"),
        "peg": rec.get("peg"),
        "float_market_cap": rec.get("float_market_cap"),
        "listing_days": rec.get("listing_days"),
        "industry": rec.get("industry", "未知"),
        "universe_score": rec.get("universe_score"),
        "float_market_cap_score": rec.get("float_market_cap_score"),
        "amount_rank_score": rec.get("amount_rank_score"),
        "universe_rank": rec.get("universe_rank"),
        "ROE": rec.get("ROE"),
        "经营现金流/净利润": rec.get("经营现金流/净利润"),
        "扣非净利润同比": rec.get("扣非净利润同比"),
        "from_candidate_pool": bool(rec.get("from_candidate_pool", False)),
        "from_holding_only": bool(rec.get("from_holding_only", False)),
    }
    if not code or asset_type not in {"STOCK", "ETF"}:
        return None, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "基础过滤",
            "failure_reason": "代码或标的类型无效",
        }
    try:
        stdout_ctx, stderr_ctx = suppress_output()
        with stdout_ctx, stderr_ctx:
            hist = build_history_frame(asset_type, code, cfg)
    except Exception as exc:
        return None, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "历史数据",
            "failure_reason": f"{type(exc).__name__}: {exc}",
        }
    if hist.empty:
        return None, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "历史数据",
            "failure_reason": "历史数据为空",
        }
    if len(hist) < int(cfg["filter"]["min_history_bars"]):
        return None, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "历史数据",
            "failure_reason": "历史数据不足120个交易日",
        }
    try:
        hist = apply_spot_price_to_history(hist, rec, cfg)
    except Exception as exc:
        return None, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "最新价格校验",
            "failure_reason": f"{type(exc).__name__}: {exc}",
        }
    metrics = compute_indicators(hist)
    row = {**base, **metrics}
    reasons = evaluate_base_filters(row, cfg)
    if reasons:
        row["base_failed"] = True
        return row, {
            "代码": code,
            "名称": name,
            "asset_type": asset_type,
            "stage": "基础过滤",
            "failure_reason": "；".join(reasons),
        }
    row["base_failed"] = False
    return row, None


def run_analysis(cfg: Dict[str, Any], holdings: Optional[pd.DataFrame] = None) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    global HISTORY_CACHE_READ_ENABLED
    HISTORY_CACHE_READ_ENABLED = bool(cfg["data"].get("cache_history", True))
    if holdings is None:
        holdings = safe_read_csv(HOLDINGS_PATH)
    candidate_seed, seed_failures = build_analysis_universe(cfg, holdings)
    candidate_seed, early_failures = early_filter_rows(candidate_seed, cfg)
    failures = seed_failures + early_failures
    if candidate_seed.empty:
        return pd.DataFrame(), holdings, pd.DataFrame(failures)
    print(f"初筛后需要获取历史行情：{len(candidate_seed)} 只")
    results: List[Dict[str, Any]] = []
    extra_failures: List[Dict[str, Any]] = []
    max_workers = min(int(cfg["data"].get("max_workers", 12)), max(1, len(candidate_seed)))
    with ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_map = {
            executor.submit(fetch_and_analyze_one, rec, cfg): rec
            for rec in candidate_seed.to_dict("records")
        }
        for future in as_completed(future_map):
            row, failure = future.result()
            if row is not None:
                results.append(row)
            if failure is not None:
                extra_failures.append(failure)
    failures.extend(extra_failures)
    analyzed = pd.DataFrame(results)
    if analyzed.empty:
        return analyzed, holdings, pd.DataFrame(failures)
    market = calc_market_vitality(analyzed, cfg)
    scored_input = analyzed[~analyzed["base_failed"].fillna(False)].copy()
    if scored_input.empty:
        scored = pd.DataFrame()
        scored.attrs["market_vitality"] = market
        return scored, holdings, pd.DataFrame(failures)
    scored = score_candidates(scored_input, cfg)
    scored.attrs["market_vitality"] = market
    target_position = float(market["target_position"])
    top_n = int(cfg.get("rebalance", {}).get("buy_top_n", cfg["strategy"]["top_n"]))
    scored["target_weight"] = target_position / float(top_n)
    holding_keys = set(zip(holdings["asset_type"].astype(str), holdings["code"].astype(str))) if not holdings.empty else set()
    scored["key"] = scored["asset_type"].astype(str) + "_" + scored["code"].astype(str)
    scored["is_holding"] = scored.apply(lambda r: (r["asset_type"], r["code"]) in holding_keys, axis=1)
    scored["action"] = scored.apply(lambda r: decide_candidate_action(r, top_n)[0], axis=1)
    scored["reason"] = scored.apply(lambda r: decide_candidate_action(r, top_n)[1], axis=1)
    scored["asset_type"] = scored["asset_type"].map(chinese_asset_type)
    scored["above_ma60"] = scored["above_ma60"].map(chinese_bool)
    scored["above_ma120"] = scored["above_ma120"].map(chinese_bool)
    scored["ma20_above_ma60"] = scored["ma20_above_ma60"].map(chinese_bool)
    scored["is_holding"] = scored["is_holding"].map(chinese_bool)
    return scored, holdings, pd.DataFrame(failures)


def build_holdings_check(scored: pd.DataFrame, holdings: pd.DataFrame, cfg: Dict[str, Any], failures: pd.DataFrame) -> pd.DataFrame:
    holdings = security_holdings(holdings)
    if holdings.empty:
        return pd.DataFrame(
            columns=[
                "code",
                "name",
                "asset_type",
                "shares",
                "cost_price",
                "latest_close",
                "pnl_amount",
                "pnl_pct",
                "rank",
                "pe_ttm",
                "valuation_score",
                "score_momentum_trend",
                "score_liquidity",
                "score_risk",
                "action",
                "reason",
                "trend_ok",
            ]
        )
    lookup = {}
    if not scored.empty:
        for _, row in scored.iterrows():
            lookup[(str(row["code"]), str(row["asset_type"]).replace("股票", "STOCK").replace("ETF", "ETF"))] = row.to_dict()
    failure_lookup = {}
    if not failures.empty:
        for _, row in failures.iterrows():
            key = (normalize_code(row.get("代码")), infer_asset_type(normalize_code(row.get("代码")), row.get("asset_type")))
            failure_lookup[key] = row.get("failure_reason", "")
    rows = []
    for _, hold in holdings.iterrows():
        code = str(hold["code"])
        asset_type = str(hold["asset_type"])
        key = (code, asset_type)
        matched = lookup.get(key)
        if matched is None:
            matched = {
                "code": code,
                "name": hold.get("name", ""),
                "asset_type": asset_type,
                "latest_close": np.nan,
                "ma20": np.nan,
                "ma60": np.nan,
                "ma120": np.nan,
                "ret20": np.nan,
                "ret60": np.nan,
                "ret120": np.nan,
                "vol60": np.nan,
                "mdd60": np.nan,
                "avg_amount_20": np.nan,
                "avg_amount_60": np.nan,
                "amount_ratio": np.nan,
                "score_momentum": np.nan,
                "score_trend": np.nan,
                "score_volume": np.nan,
                "score_liquidity": np.nan,
                "score_risk": np.nan,
                "score_momentum_trend": np.nan,
                "valuation_score": np.nan,
                "pe_ttm": np.nan,
                "total_score": np.nan,
                "rank": np.nan,
                "trend_ok": False,
                "base_failed": True,
            }
        latest_close = to_number(matched.get("latest_close"))
        cost_price = to_number(hold.get("cost_price"))
        shares = to_number(hold.get("shares"))
        pnl_pct = (latest_close / cost_price - 1.0) if pd.notna(latest_close) and pd.notna(cost_price) and cost_price > 0 else np.nan
        pnl_amount = (latest_close - cost_price) * shares if pd.notna(latest_close) and pd.notna(cost_price) and pd.notna(shares) else np.nan
        row = {
            "code": code,
            "name": clean_name(matched.get("name", hold.get("name", ""))),
            "asset_type": asset_type,
            "shares": shares,
            "cost_price": cost_price,
            "latest_close": latest_close,
            "ma20": to_number(matched.get("ma20")),
            "ma60": to_number(matched.get("ma60")),
            "ma120": to_number(matched.get("ma120")),
            "ret20": to_number(matched.get("ret20")),
            "ret60": to_number(matched.get("ret60")),
            "ret120": to_number(matched.get("ret120")),
            "vol60": to_number(matched.get("vol60")),
            "mdd60": to_number(matched.get("mdd60")),
            "avg_amount_20": to_number(matched.get("avg_amount_20")),
            "avg_amount_60": to_number(matched.get("avg_amount_60")),
            "amount_ratio": to_number(matched.get("amount_ratio")),
            "above_ma60": matched.get("above_ma60", False),
            "above_ma120": matched.get("above_ma120", False),
            "ma20_above_ma60": matched.get("ma20_above_ma60", False),
            "trend_ok": matched.get("trend_ok", False),
            "score_momentum": to_number(matched.get("score_momentum")),
            "score_trend": to_number(matched.get("score_trend")),
            "score_volume": to_number(matched.get("score_volume")),
            "score_liquidity": to_number(matched.get("score_liquidity")),
            "score_risk": to_number(matched.get("score_risk")),
            "score_momentum_trend": to_number(matched.get("score_momentum_trend")),
            "valuation_score": to_number(matched.get("valuation_score")),
            "pe_ttm": to_number(matched.get("pe_ttm")),
            "total_score": to_number(matched.get("total_score")),
            "rank": to_int_or_none(matched.get("rank")),
            "pnl_pct": pnl_pct,
            "pnl_amount": pnl_amount,
            "base_failed": bool(matched.get("base_failed", False)),
        }
        action, reason = decide_hold_action(row, cfg)
        row["action"] = action
        row["reason"] = reason
        rows.append(row)
    out = pd.DataFrame(rows)
    order = {"建议卖出": 0, "建议减仓": 1, "观察": 2, "继续持有": 3}
    out["action_rank"] = out["action"].map(order).fillna(9)
    out = out.sort_values(["action_rank", "rank", "pnl_pct"], ascending=[True, True, True]).drop(columns=["action_rank"]).reset_index(drop=True)
    out["asset_type"] = out["asset_type"].map(chinese_asset_type)
    out["above_ma60"] = out["above_ma60"].map(chinese_bool)
    out["above_ma120"] = out["above_ma120"].map(chinese_bool)
    out["ma20_above_ma60"] = out["ma20_above_ma60"].map(chinese_bool)
    out["trend_ok"] = out["trend_ok"].map(chinese_bool)
    return out


def build_action_lists(scored: pd.DataFrame, holdings_check: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    if scored.empty:
        return pd.DataFrame(), pd.DataFrame(), pd.DataFrame()
    top_candidates = scored.head(int(scored.shape[0] and min(10, len(scored)))).copy()
    if not top_candidates.empty:
        top_candidates["action"] = top_candidates.apply(lambda r: decide_candidate_action(r, 10)[0], axis=1)
        top_candidates["reason"] = top_candidates.apply(lambda r: decide_candidate_action(r, 10)[1], axis=1)
    buy_rows = top_candidates[[
        "rank",
        "code",
        "name",
        "asset_type",
        "latest_close",
        "total_score",
        "universe_score",
        "valuation_score",
        "pe_ttm",
        "score_momentum_trend",
        "score_liquidity",
        "score_risk",
        "target_weight",
        "action",
        "reason",
    ]].copy() if not top_candidates.empty else pd.DataFrame()

    reduce_rows = holdings_check[holdings_check["action"] == "建议减仓"].copy()
    sell_rows = holdings_check[holdings_check["action"] == "建议卖出"].copy()
    for df in [reduce_rows, sell_rows]:
        if not df.empty:
            df["action"] = df["action"].map(lambda x: x)
    return buy_rows, reduce_rows, sell_rows


def write_summary_sheet(ws, summary: Dict[str, Any], holdings_count: int, stock_candidate_count: int, etf_candidate_count: int, buy_count: int, reduce_count: int, sell_count: int) -> None:
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "A股每周量化调仓系统摘要"
    title.fill = TITLE_FILL
    title.font = Font(name="Microsoft YaHei", color="FFFFFFFF", bold=True, size=15)
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    state_fill = CARD_FILL_MAP.get("default")
    market_state = summary["state"]
    if market_state == "弱市场":
        state_fill = CARD_FILL_MAP["weak"]
    elif market_state == "普通市场":
        state_fill = CARD_FILL_MAP["normal"]
    elif market_state == "强市场":
        state_fill = CARD_FILL_MAP["strong"]
    elif market_state == "活跃市场":
        state_fill = CARD_FILL_MAP["hot"]
    add_summary_card(ws, 3, 1, "运行日期", current_date_obj().strftime("%Y-%m-%d"), "周度运行", CARD_FILL_MAP["default"])
    add_summary_card(ws, 3, 3, "市场活力分", f"{int(summary['score'])}/5", "候选池综合活力", state_fill)
    add_summary_card(ws, 3, 5, "市场状态", market_state, "根据活力分自动判断", state_fill)
    add_summary_card(ws, 3, 7, "建议权益仓位", f"{summary['target_position']:.0%}", "仓位建议", state_fill)
    add_summary_card(ws, 7, 1, "单只目标仓位", f"{summary['target_position'] / float(max(1, int(summary['top_n']))):.0%}", f"按 {int(summary['top_n'])} 只目标仓位估算", CARD_FILL_MAP["default"])
    add_summary_card(ws, 7, 3, "候选标的数量", str(summary["candidate_count"]), f"股票 {stock_candidate_count} 只 / ETF {etf_candidate_count} 只", CARD_FILL_MAP["default"])
    add_summary_card(ws, 7, 5, "当前持仓数量", str(holdings_count), "读取 holdings.csv", CARD_FILL_MAP["default"])
    add_summary_card(ws, 7, 7, "操作指令", "见下方说明", "先卖出，再减仓，最后买入", state_fill)

    left_pairs = [
        ("股票候选数量", str(stock_candidate_count)),
        ("ETF候选数量", str(etf_candidate_count)),
        ("实际买入数量", str(buy_count)),
        ("建议减仓数量", str(reduce_count)),
        ("建议卖出数量", str(sell_count)),
        ("账户总资产", f"{float(summary.get('total_asset', summary.get('reference_capital', 0))):,.0f} 元"),
        ("当前持仓市值", f"{float(summary.get('current_equity_value', 0)):,.0f} 元"),
        ("持仓成本", f"{float(summary.get('holding_cost', 0)):,.0f} 元"),
        ("当前浮动盈亏", f"{float(summary.get('unrealized_pnl', 0)):,.0f} 元"),
        ("当前浮动盈亏率", f"{float(summary.get('unrealized_pnl_pct', 0)):.2%}"),
        ("已实现盈亏", f"{float(summary.get('realized_pnl', 0)):,.0f} 元"),
        ("累计盈亏", f"{float(summary.get('total_pnl', 0)):,.0f} 元"),
        ("剩余现金", f"{float(summary.get('cash_balance', 0)):,.0f} 元"),
        ("当前权益仓位", f"{float(summary.get('current_position', 0)):.0%}"),
        ("实际买入预算", f"{float(summary.get('actual_buy_budget', float(summary.get('reference_capital', 0)) * float(summary.get('target_position', 0)))):,.0f} 元"),
        ("预留现金", f"{float(summary.get('cash_buffer', 0)):,.0f} 元"),
    ]
    write_key_value_table(ws, 12, 1, "统计明细", left_pairs, width=24)
    right_pairs = [
        ("市场活力分", f"{int(summary['score'])}/5"),
        ("市场状态", market_state),
        ("建议权益仓位", f"{summary['target_position']:.0%}"),
        ("单只目标仓位", f"{summary['target_position'] / float(max(1, int(summary['top_n']))):.0%}"),
        ("交易记录状态", summary.get("trade_record_note", "未读取 trades.csv")),
        ("操作指令", summary["core_recommendation"]),
    ]
    row_end = write_key_value_table(ws, 12, 5, "仓位建议", right_pairs, width=30)
    ws.merge_cells(start_row=row_end + 1, start_column=1, end_row=row_end + 3, end_column=8)
    note = ws.cell(row=row_end + 1, column=1, value=summary["core_recommendation"])
    note.fill = PatternFill("solid", fgColor=state_fill)
    note.font = Font(name="Microsoft YaHei", color="FF1F1F1F", bold=True, size=11)
    note.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    note.border = THIN_BORDER
    for r in range(row_end + 1, row_end + 4):
        for c in range(1, 9):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=state_fill)
            ws.cell(row=r, column=c).border = THIN_BORDER
    ws.row_dimensions[row_end + 1].height = 36
    ws.row_dimensions[row_end + 2].height = 36
    ws.row_dimensions[row_end + 3].height = 36
    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["H"].width = 34
    ws.sheet_view.showGridLines = False


def prepare_output_dfs(
    scored: pd.DataFrame,
    holdings: pd.DataFrame,
    holdings_check: pd.DataFrame,
    failures: pd.DataFrame,
    cfg: Dict[str, Any],
    account_summary: Dict[str, float],
) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    top_n = int(cfg.get("rebalance", {}).get("buy_top_n", cfg["strategy"]["top_n"]))
    lot_size = 100
    existing_equity_value = float(account_summary.get("current_equity_value", 0.0))
    total_asset = float(account_summary.get("total_asset", 0.0))
    available_cash = float(account_summary.get("cash_balance", 0.0))
    scored = scored.copy()
    if scored.empty:
        scored_export = pd.DataFrame(
            columns=[
                "排名",
                "代码",
                "名称",
                "标的类型",
                "当前持仓",
                "最新价",
                "20日均线",
                "60日均线",
                "120日均线",
                "近20日涨跌幅",
                "近60日涨跌幅",
                "近120日涨跌幅",
                "近60日年化波动率",
                "近60日最大回撤",
                "近20日日均成交额",
                "近60日日均成交额",
                "成交额放大倍数",
                "是否站上60日线",
                "是否站上120日线",
                "20日线是否高于60日线",
                "最新价/60日均线",
                "20日均线/60日均线",
                "最新价/120日均线",
                "动量得分",
                "趋势得分",
                "成交活跃得分",
                "流动性得分",
                "风险控制得分",
                "综合评分",
                "建议目标仓位",
                "建议买入股数",
                "建议交易金额",
                "建议执行",
                "操作建议",
                "建议理由",
            ]
        )
        buy_rows = pd.DataFrame()
        reduce_rows = pd.DataFrame()
        sell_rows = pd.DataFrame()
    else:
        scored["is_holding"] = scored["is_holding"].map(chinese_bool)
        scored["asset_type"] = scored["asset_type"].map(lambda x: x if x in {"股票", "ETF"} else chinese_asset_type(x))
        scored["action"], scored["reason"] = zip(*scored.apply(lambda r: decide_candidate_action(r, top_n), axis=1))
        scored["target_weight"] = scored["target_weight"].astype(float)
        target_position = float(scored["target_weight"].iloc[0]) * float(top_n) if not scored.empty else float(cfg["strategy"]["target_position_normal"])
        buy_rows, buy_lookup = build_actual_buy_plan(
            scored,
            total_asset=total_asset,
            target_position=target_position,
            top_n=top_n,
            existing_equity_value=existing_equity_value,
            available_cash=available_cash,
            lot_size=lot_size,
        )
        scored["buy_shares"] = scored["code"].map(lambda code: int(buy_lookup.get(str(code), {}).get("buy_shares", 0)))
        scored["trade_value"] = scored["code"].map(lambda code: buy_lookup.get(str(code), {}).get("trade_value", np.nan))
        scored["trade_instruction"] = scored["code"].map(lambda code: buy_lookup.get(str(code), {}).get("trade_instruction", "不操作"))
        scored_export = export_columns(
            scored,
            [
                "rank",
                "code",
                "name",
                "asset_type",
                "is_holding",
                "latest_close",
                "ma20",
                "ma60",
                "ma120",
                "ret20",
                "ret60",
                "ret120",
                "vol60",
                "mdd60",
                "avg_amount_20",
                "avg_amount_60",
                "amount_ratio",
                "above_ma60",
                "above_ma120",
                "ma20_above_ma60",
                "price_ma60_ratio",
                "ma20_ma60_ratio",
                "price_ma120_ratio",
                "universe_score",
                "universe_rank",
                "float_market_cap_score",
                "amount_rank_score",
                "float_market_cap",
                "industry",
                "ROE",
                "经营现金流/净利润",
                "扣非净利润同比",
                "pe_ttm",
                "valuation_score",
                "score_momentum_trend",
                "score_liquidity",
                "score_risk",
                "total_score",
                "入选原因",
                "风险提示",
                "score_momentum",
                "score_trend",
                "score_volume",
                "target_weight",
                "buy_shares",
                "trade_value",
                "trade_instruction",
                "action",
                "reason",
            ],
        )
        _, reduce_rows, sell_rows = build_action_lists(scored, holdings_check)
        if not buy_rows.empty:
            buy_rows = export_columns(
                buy_rows,
                [
                    "rank",
                    "code",
                    "name",
                    "asset_type",
                    "is_holding",
                    "latest_close",
                    "total_score",
                    "universe_score",
                    "valuation_score",
                    "pe_ttm",
                    "score_momentum_trend",
                    "score_liquidity",
                    "score_risk",
                    "target_weight",
                    "buy_shares",
                    "trade_value",
                    "trade_instruction",
                    "action",
                    "reason",
                ],
            )
        if not reduce_rows.empty:
            reduce_rows["sell_shares"], reduce_rows["trade_value"], reduce_rows["trade_instruction"] = zip(
                *reduce_rows.apply(lambda r: recommend_reduce_trade(r, lot_size), axis=1)
            )
            reduce_rows = export_columns(
                reduce_rows,
                [
                    "action",
                    "code",
                    "name",
                    "asset_type",
                    "shares",
                    "cost_price",
                    "latest_close",
                    "pnl_amount",
                    "pnl_pct",
                    "rank",
                    "total_score",
                    "sell_shares",
                    "trade_value",
                    "trade_instruction",
                    "reason",
                ],
            )
        if not sell_rows.empty:
            sell_rows["sell_shares"], sell_rows["trade_value"], sell_rows["trade_instruction"] = zip(
                *sell_rows.apply(lambda r: recommend_sell_all_trade(r), axis=1)
            )
            sell_rows = export_columns(
                sell_rows,
                [
                    "action",
                    "code",
                    "name",
                    "asset_type",
                    "shares",
                    "cost_price",
                    "latest_close",
                    "pnl_amount",
                    "pnl_pct",
                    "rank",
                    "total_score",
                    "sell_shares",
                    "trade_value",
                    "trade_instruction",
                    "reason",
                ],
            )
    if not holdings_check.empty:
        holdings_check = holdings_check.copy()
        holdings_check["sell_shares"], holdings_check["trade_value"], holdings_check["trade_instruction"] = zip(
            *holdings_check.apply(
                lambda r: recommend_reduce_trade(r, lot_size) if r.get("action") == "建议减仓" else recommend_sell_all_trade(r) if r.get("action") == "建议卖出" else (0, np.nan, "不操作"),
                axis=1,
            )
        )
        holdings_export = export_columns(
            holdings_check,
            [
                "action",
                "code",
                "name",
                "asset_type",
                "shares",
                "cost_price",
                "latest_close",
                "pnl_pct",
                "rank",
                "score_momentum",
                "score_trend",
                "score_volume",
                "score_liquidity",
                "score_risk",
                "ret20",
                "ret60",
                "ret120",
                "vol60",
                "mdd60",
                "avg_amount_20",
                "amount_ratio",
                "above_ma60",
                "above_ma120",
                "ma20_above_ma60",
                "total_score",
                "sell_shares",
                "trade_value",
                "trade_instruction",
                "reason",
            ],
        )
    else:
        holdings_export = pd.DataFrame(
            columns=[
                "操作建议",
                "代码",
                "名称",
                "标的类型",
                "持仓数量",
                "成本价",
                "最新价",
                "浮动盈亏金额",
                "浮动盈亏率",
                "排名",
                "动量得分",
                "趋势得分",
                "成交活跃得分",
                "流动性得分",
                "风险控制得分",
                "近20日涨跌幅",
                "近60日涨跌幅",
                "近120日涨跌幅",
                "近60日年化波动率",
                "近60日最大回撤",
                "近20日日均成交额",
                "成交额放大倍数",
                "是否站上60日线",
                "是否站上120日线",
                "20日线是否高于60日线",
                "综合评分",
                "建议卖出股数",
                "建议交易金额",
                "建议执行",
                "建议理由",
            ]
        )
    if not failures.empty:
        failures_export = export_columns(failures, ["code", "name", "asset_type", "stage", "failure_reason"])
    else:
        failures_export = pd.DataFrame(columns=["代码", "名称", "标的类型", "失败阶段", "失败原因"])
    return scored_export, holdings_export, buy_rows, reduce_rows, sell_rows, failures_export


def auto_fill_header_sheet(ws, df: pd.DataFrame, action_col_name: Optional[str] = None) -> None:
    if df.empty:
        return
    headers = list(df.columns)
    action_col_idx = headers.index(action_col_name) + 1 if action_col_name and action_col_name in headers else None
    for r_idx in range(2, len(df) + 2):
        action = ""
        if action_col_idx is not None:
            action = str(ws.cell(row=r_idx, column=action_col_idx).value or "")
        fill = PatternFill("solid", fgColor=ACTION_FILL_MAP.get(action, "FFFFFFFF")) if action in ACTION_FILL_MAP else None
        if fill is not None:
            for c in range(1, len(headers) + 1):
                ws.cell(row=r_idx, column=c).fill = fill


def create_workbook(
    summary: Dict[str, Any],
    scored_export: pd.DataFrame,
    holdings_export: pd.DataFrame,
    buy_rows: pd.DataFrame,
    reduce_rows: pd.DataFrame,
    sell_rows: pd.DataFrame,
    failures_export: pd.DataFrame,
    trade_pnl_export: pd.DataFrame,
    holdings_count: int,
    stock_candidate_count: int,
    etf_candidate_count: int,
) -> Workbook:
    wb = Workbook()
    default = wb.active
    wb.remove(default)

    # 调仓建议
    ws = wb.create_sheet("调仓建议")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:H1")
    title = ws["A1"]
    title.value = "本周调仓建议"
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28
    summary_pairs = [
        ("运行日期", current_date_obj().strftime("%Y-%m-%d")),
        ("市场活力分", f"{int(summary['score'])}/5"),
        ("市场状态", summary["state"]),
        ("建议权益仓位", f"{summary['target_position']:.0%}"),
        ("单只目标仓位", f"{summary['target_position'] / float(max(1, int(summary['top_n']))):.0%}"),
        ("候选标的数量", str(summary["candidate_count"])),
        ("当前持仓数量", str(holdings_count)),
    ]
    row = 3
    for label, value in summary_pairs:
        ws.cell(row=row, column=1, value=label).font = BODY_BOLD_FONT
        ws.cell(row=row, column=2, value=value).font = BODY_FONT
        for c in range(1, 3):
            ws.cell(row=row, column=c).border = THIN_BORDER
            ws.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFF8FAFC")
        row += 1
    ws.merge_cells(start_row=3, start_column=4, end_row=9, end_column=8)
    core = ws.cell(row=3, column=4, value=summary["core_recommendation"])
    core.fill = PatternFill("solid", fgColor=CARD_FILL_MAP.get("default"))
    core.font = Font(name="Microsoft YaHei", bold=True, size=11)
    core.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    core.border = THIN_BORDER
    for r in range(3, 10):
        for c in range(4, 9):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=CARD_FILL_MAP.get("default"))

    cur_row = 11
    if not buy_rows.empty:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=max(1, buy_rows.shape[1]))
        ws.cell(row=cur_row, column=1, value="可新开仓").fill = PatternFill("solid", fgColor="FFD9EAF7")
        ws.cell(row=cur_row, column=1).font = BODY_BOLD_FONT
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=cur_row, column=1).border = THIN_BORDER
        write_dataframe_sheet(
            ws,
            buy_rows,
            title=None,
            action_col="操作建议",
            wrap_cols=["建议理由"],
            extra_widths={"建议理由": 44},
            start_row=cur_row + 1,
            apply_filter=False,
        )
        cur_row = cur_row + len(buy_rows) + 3
    if not reduce_rows.empty:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=max(1, reduce_rows.shape[1]))
        ws.cell(row=cur_row, column=1, value="建议减仓").fill = PatternFill("solid", fgColor="FFFCE5CD")
        ws.cell(row=cur_row, column=1).font = BODY_BOLD_FONT
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=cur_row, column=1).border = THIN_BORDER
        write_dataframe_sheet(
            ws,
            reduce_rows,
            title=None,
            action_col="操作建议",
            wrap_cols=["建议理由"],
            extra_widths={"建议理由": 44},
            start_row=cur_row + 1,
            apply_filter=False,
        )
        cur_row = cur_row + len(reduce_rows) + 3
    if not sell_rows.empty:
        ws.merge_cells(start_row=cur_row, start_column=1, end_row=cur_row, end_column=max(1, sell_rows.shape[1]))
        ws.cell(row=cur_row, column=1, value="建议卖出").fill = PatternFill("solid", fgColor="FFF4CCCC")
        ws.cell(row=cur_row, column=1).font = BODY_BOLD_FONT
        ws.cell(row=cur_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
        ws.cell(row=cur_row, column=1).border = THIN_BORDER
        write_dataframe_sheet(
            ws,
            sell_rows,
            title=None,
            action_col="操作建议",
            wrap_cols=["建议理由"],
            extra_widths={"建议理由": 44},
            start_row=cur_row + 1,
            apply_filter=False,
        )

    for c in range(1, 9):
        ws.column_dimensions[get_column_letter(c)].width = 16
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["H"].width = 40

    # 系统摘要
    ws = wb.create_sheet("系统摘要")
    write_summary_sheet(
        ws,
        summary,
        holdings_count=holdings_count,
        stock_candidate_count=stock_candidate_count,
        etf_candidate_count=etf_candidate_count,
        buy_count=len(buy_rows),
        reduce_count=len(reduce_rows),
        sell_count=len(sell_rows),
    )

    # 实际买入计划
    ws = wb.create_sheet("实际买入计划")
    ws.sheet_view.showGridLines = False
    ws.merge_cells("A1:R1")
    title = ws["A1"]
    title.value = "实际买入计划（按账户现金和目标仓位缺口生成）"
    title.fill = TITLE_FILL
    title.font = TITLE_FONT
    title.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24
    actual_buy_budget = float(summary.get("actual_buy_budget", float(summary.get("reference_capital", 0)) * float(summary.get("target_position", 0))))
    cash_buffer = float(summary.get("cash_buffer", max(0.0, float(summary.get("cash_balance", 0)) - actual_buy_budget)))
    note_pairs = [
        ("账户总资产", f"{float(summary.get('total_asset', summary.get('reference_capital', 0))):,.0f} 元"),
        ("当前持仓市值", f"{float(summary.get('current_equity_value', 0)):,.0f} 元"),
        ("当前浮动盈亏", f"{float(summary.get('unrealized_pnl', 0)):,.0f} 元"),
        ("已实现盈亏", f"{float(summary.get('realized_pnl', 0)):,.0f} 元"),
        ("累计盈亏", f"{float(summary.get('total_pnl', 0)):,.0f} 元"),
        ("剩余现金", f"{float(summary.get('cash_balance', 0)):,.0f} 元"),
        ("当前权益仓位", f"{float(summary.get('current_position', 0)):.0%}"),
        ("建议权益仓位", f"{summary['target_position']:.0%}"),
        ("实际买入预算", f"{actual_buy_budget:,.0f} 元"),
        ("预留现金", f"{cash_buffer:,.0f} 元"),
        ("实际买入数量", str(len(buy_rows))),
    ]
    for offset, (label, value) in enumerate(note_pairs, start=3):
        ws.cell(row=offset, column=1, value=label).font = BODY_BOLD_FONT
        ws.cell(row=offset, column=1).border = THIN_BORDER
        ws.cell(row=offset, column=1).fill = PatternFill("solid", fgColor="FFF8FAFC")
        ws.cell(row=offset, column=2, value=value).font = BODY_FONT
        ws.cell(row=offset, column=2).border = THIN_BORDER
        ws.cell(row=offset, column=2).fill = PatternFill("solid", fgColor="FFF8FAFC")
        for c in range(3, 19):
            ws.cell(row=offset, column=c).border = THIN_BORDER
            ws.cell(row=offset, column=c).fill = PatternFill("solid", fgColor="FFF8FAFC")
    write_dataframe_sheet(
        ws,
        buy_rows,
        title=None,
        action_col="操作建议",
        wrap_cols=["建议理由"],
        extra_widths={"建议理由": 46, "名称": 14},
        start_row=3 + len(note_pairs) + 2,
    )

    # 候选买入Top10
    ws = wb.create_sheet("候选买入Top10")
    if not scored_export.empty:
        top10 = scored_export.head(10).copy()
    else:
        top10 = scored_export.copy()
    write_dataframe_sheet(
        ws,
        top10,
        title="候选买入Top10（排名预览）",
        action_col="操作建议",
        wrap_cols=["建议理由"],
        extra_widths={"建议理由": 44, "名称": 14, "标的类型": 10},
    )

    # 当前持仓检查
    ws = wb.create_sheet("当前持仓检查")
    write_dataframe_sheet(
        ws,
        holdings_export,
        title=None,
        action_col="操作建议",
        wrap_cols=["建议理由"],
        extra_widths={"建议理由": 46, "名称": 14},
    )

    # 交易记录盈亏
    ws = wb.create_sheet("交易记录盈亏")
    write_dataframe_sheet(
        ws,
        trade_pnl_export,
        title=None,
        action_col=None,
        wrap_cols=["备注"],
        extra_widths={"备注": 30, "名称": 14},
    )

    # 全部评分
    ws = wb.create_sheet("全部评分")
    write_dataframe_sheet(
        ws,
        scored_export,
        title=None,
        action_col="操作建议",
        wrap_cols=["建议理由"],
        extra_widths={"建议理由": 44, "名称": 14, "标的类型": 10},
    )

    # 失败列表
    ws = wb.create_sheet("失败列表")
    write_dataframe_sheet(
        ws,
        failures_export,
        title=None,
        action_col=None,
        wrap_cols=["失败原因"],
        extra_widths={"失败原因": 60, "失败阶段": 12, "名称": 14},
    )

    return wb


def main() -> None:
    reset_status_steps()
    output_path: Optional[Path] = None
    scored_empty = False
    cfg = load_config()
    add_status_step("读取配置")
    reference_capital = float(cfg.get("strategy", {}).get("reference_capital", 20000))
    base_holdings = safe_read_csv(HOLDINGS_PATH)
    trade_records_path = resolve_trade_records_path()
    trades = safe_read_trade_records(trade_records_path)
    add_status_step("读取持仓和交易记录")
    holdings_loaded = apply_trade_records_to_holdings(base_holdings, trades)
    try:
        holdings_loaded.to_csv(AUTO_HOLDINGS_PATH, index=False, encoding="utf-8-sig")
    except Exception:
        pass
    add_status_step("滚动生成当前持仓")
    scored, holdings_loaded, failures = run_analysis(cfg, holdings_loaded)
    add_status_step("获取行情并完成候选打分")
    holdings_check = build_holdings_check(scored, holdings_loaded, cfg, failures)
    add_status_step("生成当前持仓检查")
    account_summary = build_account_summary(holdings_loaded, holdings_check, reference_capital)
    trade_pnl_summary, trade_pnl_export, _, _ = build_realized_pnl_from_trades(trades, base_holdings)
    add_status_step("统计账户资产和交易盈亏")
    realized_pnl = float(trade_pnl_summary.get("realized_pnl", 0.0))
    unrealized_pnl = float(account_summary.get("unrealized_pnl", 0.0))
    trade_record_note = "未提供交割单，仅按 holdings.csv 统计当前持仓和现金。"
    if trade_records_path is not None and trade_records_path.exists():
        trade_record_note = f"已读取 {trade_records_path.name}，共 {len(trades)} 条有效交易记录，已自动滚动持仓和现金。"
        if float(trade_pnl_summary.get("unmatched_sell_shares", 0.0)) > 0:
            trade_record_note += " 有卖出记录缺少对应买入成本，相关已实现盈亏无法完整匹配。"
    if trade_pnl_export.empty:
        trade_pnl_export = pd.DataFrame(
            columns=["日期", "方向", "代码", "名称", "标的类型", "成交数量", "成交价格", "成交金额", "卖出匹配成本", "已实现盈亏", "备注"]
        )
    total_asset = float(account_summary["total_asset"])
    holding_count = len(security_holdings(holdings_loaded))
    if not scored.empty:
        market = scored.attrs.get("market_vitality") or calc_market_vitality(scored, cfg)
        top_n = int(cfg.get("rebalance", {}).get("buy_top_n", cfg["strategy"]["top_n"]))
        target_position = float(market["target_position"])
        buy_candidates = scored.head(min(top_n, len(scored)))
        buy_count = int(((buy_candidates["is_holding"] == "否") | (buy_candidates["is_holding"] == False)).sum()) if not buy_candidates.empty else 0
        reduce_count = int((holdings_check["action"] == "建议减仓").sum()) if not holdings_check.empty else 0
        sell_count = int((holdings_check["action"] == "建议卖出").sum()) if not holdings_check.empty else 0
        core_recommendation = build_core_recommendation(market, buy_count, reduce_count, sell_count, top_n, total_asset)
        stock_candidate_count = int((scored["asset_type"] == "股票").sum())
        etf_candidate_count = int((scored["asset_type"] == "ETF").sum())
        candidate_count = len(scored)
    else:
        market = {
            "score": 0,
            "state": "无有效候选",
            "target_position": float(cfg["strategy"]["target_position_weak"]),
        }
        top_n = int(cfg.get("rebalance", {}).get("buy_top_n", cfg["strategy"]["top_n"]))
        target_position = float(market["target_position"])
        buy_count = 0
        reduce_count = int((holdings_check["action"] == "建议减仓").sum()) if not holdings_check.empty else 0
        sell_count = int((holdings_check["action"] == "建议卖出").sum()) if not holdings_check.empty else 0
        core_recommendation = "没有任何标的通过基础过滤，无法形成有效调仓建议，请检查网络、行情源或持仓文件。"
        stock_candidate_count = 0
        etf_candidate_count = 0
        candidate_count = 0
    summary = {
        "score": market["score"],
        "state": market["state"],
        "target_position": target_position,
        "candidate_count": candidate_count,
        "top_n": top_n,
        "core_recommendation": core_recommendation,
        "reference_capital": reference_capital,
        "realized_pnl": realized_pnl,
        "total_pnl": realized_pnl + unrealized_pnl,
        "trade_record_note": trade_record_note,
        **account_summary,
    }

    scored_export, holdings_export, buy_rows, reduce_rows, sell_rows, failures_export = prepare_output_dfs(
        scored, holdings_loaded, holdings_check, failures, cfg, account_summary
    )
    add_status_step("整理报告数据表")

    actual_buy_count = len(buy_rows)
    current_equity_value = float(account_summary["current_equity_value"])
    cash = float(account_summary["cash_balance"])
    target_equity_value = total_asset * float(target_position)
    position_gap = max(0.0, target_equity_value - current_equity_value)
    actual_buy_budget = min(position_gap, cash)
    summary["actual_buy_budget"] = actual_buy_budget
    summary["cash_buffer"] = max(0.0, cash - actual_buy_budget)
    if not scored.empty:
        core_recommendation = build_core_recommendation(market, actual_buy_count, reduce_count, sell_count, top_n, total_asset)
    else:
        core_recommendation = "没有任何标的通过基础过滤，无法形成有效调仓建议，请检查网络、行情源或持仓文件。"
    summary["core_recommendation"] = core_recommendation

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    output_path = OUTPUT_DIR / f"weekly_report_{current_date_text()}.xlsx"
    wb = create_workbook(
        summary,
        scored_export,
        holdings_export,
        buy_rows,
        reduce_rows,
        sell_rows,
        failures_export,
        trade_pnl_export,
        holdings_count=holding_count,
        stock_candidate_count=stock_candidate_count,
        etf_candidate_count=etf_candidate_count,
    )
    wb.save(output_path)
    if bool(cfg["data"].get("cache_history", False)):
        CACHE_DIR.mkdir(parents=True, exist_ok=True)
        successful_run_marker_path().write_text(output_path.name, encoding="utf-8")
    add_status_step("保存Excel报告")
    print(f"已生成：{output_path}")

    if scored.empty:
        scored_empty = True
        add_status_step("生成有效调仓建议", "未完成")
    else:
        add_status_step("生成有效调仓建议")
    show_status_window(
        "量化系统运行状态",
        finished=not scored_empty,
        output_path=output_path,
        error=RuntimeError("所有候选标的都失败了，已生成失败列表，但没有有效调仓建议。") if scored_empty else None,
    )


if __name__ == "__main__":
    try:
        main()
    except Exception as exc:
        add_status_step("程序异常终止", "失败")
        show_status_window("量化系统运行状态", finished=False, error=exc)
        raise
